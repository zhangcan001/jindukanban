from __future__ import annotations

import json
from io import BytesIO

from docx import Document
from fastapi.testclient import TestClient
from openpyxl import load_workbook

from app.main import app

CASES = [
    (32, 82, "\u6d88\u9632\u6279\u6b21"),
    (32, 81, "\u673a\u7535\u6279\u6b21"),
    (32, 83, "\u667a\u80fd\u5316\u6279\u6b21"),
]
WORD_TITLE = "\u516b\u3001\u8fdb\u9636\u56fe\u8868\u5206\u6790"
EXCEL_SHEETS = [
    "\u4e13\u4e1a\u8fdb\u5ea6\u5bf9\u6bd4",
    "\u697c\u5c42\u4e13\u4e1a\u77e9\u9635",
    "\u697c\u680b\u4e13\u4e1a\u77e9\u9635",
    "\u6ede\u540e\u5206\u5e03\u7edf\u8ba1",
]

summary = []
with TestClient(app) as client:
    original = client.get("/api/projects/32/reports/config").json()
    config = {
        **original,
        "include_advanced_chart_analysis": True,
        "show_data_quality_section": True,
        "show_rectification_summary": True,
        "weekly_delayed_item_limit": 18,
        "weekly_matrix_summary_limit": 7,
        "default_export_format": "docx",
        "file_name_include_project_name": True,
        "file_name_include_data_date": True,
    }
    client.put("/api/projects/32/reports/config", json=config).raise_for_status()
    try:
        no_batch = client.get("/api/projects/999999/reports/weekly-word")
        assert no_batch.status_code == 404
        for project_id, batch_id, label in CASES:
            word_preview = client.get(f"/api/projects/{project_id}/reports/preview/weekly_word", params={"batch_id": batch_id})
            word_preview.raise_for_status()
            word_items = {item["label"]: item["value"] for item in word_preview.json()["items"]}
            assert word_items["是否包含进阶图表分析"] is True
            assert word_items["主要滞后项最大条数"] == 18
            assert word_items["矩阵摘要最大条数"] == 7

            word = client.get(f"/api/projects/{project_id}/reports/weekly-word", params={"batch_id": batch_id})
            word.raise_for_status()
            word_text = "\n".join(paragraph.text for paragraph in Document(BytesIO(word.content)).paragraphs)
            assert WORD_TITLE in word_text

            excel_preview = client.get(f"/api/projects/{project_id}/reports/preview/dashboard_excel", params={"batch_id": batch_id})
            excel_preview.raise_for_status()
            excel_items = {item["label"]: item["value"] for item in excel_preview.json()["items"]}
            assert set(EXCEL_SHEETS).issubset(set(excel_items["包含 Sheet 列表"]))
            dashboard = client.get(f"/api/projects/{project_id}/reports/dashboard-export", params={"batch_id": batch_id})
            dashboard.raise_for_status()
            workbook = load_workbook(BytesIO(dashboard.content), read_only=True)
            try:
                assert set(EXCEL_SHEETS).issubset(set(workbook.sheetnames))
            finally:
                workbook.close()

            rect_preview = client.get(f"/api/projects/{project_id}/reports/preview/rectification_tracking", params={"batch_id": batch_id})
            rect_preview.raise_for_status()
            rectification = client.get(f"/api/projects/{project_id}/rectifications/export", params={"batch_id": batch_id})
            rectification_status = rectification.status_code
            if rectification_status == 200:
                rect_wb = load_workbook(BytesIO(rectification.content), read_only=True)
                try:
                    assert rect_wb.active.max_row >= 1
                finally:
                    rect_wb.close()
            else:
                assert rectification.json()["detail"]["message"] == "\u5f53\u524d\u7b5b\u9009\u6761\u4ef6\u4e0b\u6682\u65e0\u6574\u6539\u9879\u3002"

            delay_report = client.get(f"/api/projects/{project_id}/reports/delay-rectification-export", params={"batch_id": batch_id})
            delay_report.raise_for_status()

            history = client.get(f"/api/projects/{project_id}/reports/exports", params={"report_type": "weekly_word", "project_name": "\u591aSheet", "keyword": "\u5468\u62a5"})
            history.raise_for_status()
            summary.append(
                {
                    "batch": label,
                    "word_preview": True,
                    "settings_saved": True,
                    "word_export": True,
                    "dashboard_excel": True,
                    "rectification_preview": True,
                    "rectification_status": rectification_status,
                    "delay_rectification": True,
                    "history_filter_count": len(history.json()),
                }
            )
    finally:
        client.put("/api/projects/32/reports/config", json=original).raise_for_status()

print(json.dumps(summary, ensure_ascii=False, indent=2))
