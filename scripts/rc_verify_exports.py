from __future__ import annotations

import sys
from pathlib import Path

from docx import Document
from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
SMOKE = Path(sys.argv[1]).resolve() if len(sys.argv) > 1 else ROOT / ".runtime" / "rc-smoke"


def assert_true(condition: bool, message: str) -> None:
    if not condition:
        raise AssertionError(message)


dashboard = SMOKE / "dashboard.xlsx"
rectification = SMOKE / "rectification.xlsx"
rectification_tracking = SMOKE / "rectification-tracking.xlsx"
warnings = SMOKE / "warnings.xlsx"
weekly = SMOKE / "weekly.docx"

workbook = load_workbook(dashboard, read_only=True)
try:
    expected_sheets = [
        "看板总览",
        "专业进度统计",
        "楼层进度统计",
        "楼栋楼层统计",
        "滞后项清单",
        "数据质量与校验问题汇总",
        "进度分析说明",
        "整改闭环摘要",
        "整改项明细",
        "专业进度对比",
        "楼层专业矩阵",
        "楼栋专业矩阵",
        "滞后分布统计",
    ]
    missing_sheets = [sheet for sheet in expected_sheets if sheet not in workbook.sheetnames]
    assert_true(not missing_sheets, f"dashboard sheets missing: {missing_sheets}; got {workbook.sheetnames}")
    values = [
        cell
        for sheet in ("整改闭环摘要", "整改项明细", "专业进度对比", "楼层专业矩阵", "楼栋专业矩阵", "滞后分布统计")
        for row in workbook[sheet].iter_rows(values_only=True)
        for cell in row
    ]
    for text in [
        "整改闭环摘要",
        "逾期整改项",
        "整改项明细",
        "是否逾期",
        "专业进度对比",
        "楼层专业矩阵",
        "楼栋专业矩阵",
        "滞后分布统计",
        "当前筛选条件",
    ]:
        assert_true(text in values, f"dashboard rectification content missing {text}")
finally:
    workbook.close()

workbook = load_workbook(rectification, read_only=True)
try:
    values = [cell for row in workbook.active.iter_rows(values_only=True) for cell in row]
    for text in [
        "专业",
        "楼栋",
        "楼层",
        "系统",
        "施工项",
        "实际完成率",
        "计划完成率",
        "偏差",
        "滞后等级",
        "滞后说明",
        "整改建议",
        "责任人",
        "计划完成时间",
        "复查结果",
        "备注",
    ]:
        assert_true(text in values, f"rectification missing {text}")
finally:
    workbook.close()

workbook = load_workbook(rectification_tracking, read_only=True)
try:
    values = [cell for row in workbook.active.iter_rows(values_only=True) for cell in row]
    for text in ["整改跟踪表", "整改记录摘要", "责任人", "责任单位", "计划完成时间", "是否逾期", "最近更新时间"]:
        assert_true(text in values, f"rectification tracking missing {text}")
finally:
    workbook.close()

workbook = load_workbook(warnings, read_only=True)
try:
    values = [cell for row in workbook.active.iter_rows(values_only=True) for cell in row]
    for text in ["专业", "楼栋", "楼层", "系统", "施工项", "预警说明"]:
        assert_true(text in values, f"warnings export missing {text}")
finally:
    workbook.close()

document = Document(weekly)
text = "\n".join(paragraph.text for paragraph in document.paragraphs)
for expected in [
    "工程进度周报",
    "一、总体进度概况",
    "重点指标表",
    "二、分专业进度情况",
    "三、楼层进度情况",
    "四、楼栋楼层进度情况",
    "五、主要滞后项",
    "六、数据质量与校验问题",
    "七、进度分析说明",
    "整改闭环摘要",
    "本期关注事项",
]:
    assert_true(expected in text, f"weekly word missing {expected}")

print("export files verified")
