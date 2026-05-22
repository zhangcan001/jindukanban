from __future__ import annotations

from datetime import date, timedelta
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
OUTPUT = ROOT / "sample_data" / "工程进度管理系统_全功能模拟测试表_v1.xlsx"

HEADERS = [
    "WBS编码",
    "楼栋",
    "楼层",
    "专业",
    "系统",
    "工作内容",
    "单位",
    "总工程量",
    "计划完成量",
    "累计完成量",
    "本期完成量",
    "计划完成率",
    "实际完成率",
    "计划开始",
    "计划完成",
    "责任人",
    "施工单位",
    "备注",
]


def main() -> None:
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    readme = workbook.active
    readme.title = "README_使用说明"
    write_readme(readme)

    write_progress_sheet(workbook, "01_单Sheet标准进度表", base_code="S", rows=standard_rows())
    write_progress_sheet(workbook, "02_多Sheet_机电单位", base_code="M", rows=discipline_rows("机电", "A座", "给排水"))
    write_progress_sheet(workbook, "03_多Sheet_消防单位", base_code="F", rows=discipline_rows("消防", "B座", "喷淋系统"))
    write_progress_sheet(workbook, "04_多Sheet_智能化单位", base_code="I", rows=discipline_rows("智能化", "C座", "弱电系统"))
    write_progress_sheet(workbook, "05_异常数据校验", base_code="E", rows=abnormal_rows())
    write_merged_header_sheet(workbook)
    write_non_progress_field_sheet(workbook)
    write_non_progress_issue_sheet(workbook)
    write_helper_sheet(workbook)

    for sheet in workbook.worksheets:
        autosize(sheet)
        freeze_header(sheet)

    workbook.save(OUTPUT)
    print(f"generated: {OUTPUT}")


def write_readme(sheet) -> None:
    rows = [
        ["工程进度管理系统全功能自动化验收测试表"],
        ["用途", "供 scripts\\full_auto_check.bat 自动验收使用。"],
        ["正常导入", "01 / 02 / 03 / 04 应可导入并发布。"],
        ["异常校验", "05 含负数工程量，应产生 error 且不可发布。"],
        ["合并表头", "06 使用两级合并表头，验证表头解析。"],
        ["非进度", "07 / 08 不应发布。"],
        ["辅助 Sheet", "09 为计划基线说明，不应发布。"],
    ]
    for row in rows:
        sheet.append(row)
    sheet["A1"].font = Font(bold=True, size=14)


def write_progress_sheet(workbook: Workbook, title: str, base_code: str, rows: list[list[object]]) -> None:
    sheet = workbook.create_sheet(title)
    sheet.append(HEADERS)
    for row in rows:
        sheet.append(row)
    style_header(sheet)


def standard_rows() -> list[list[object]]:
    today = date.today()
    delayed_start = (today - timedelta(days=18)).isoformat()
    delayed_finish = (today + timedelta(days=6)).isoformat()
    overdue_finish = (today - timedelta(days=2)).isoformat()
    normal_start = today.isoformat()
    return [
        ["S-001", "B2", "2层", "机电", "桥架系统", "机电桥架安装", "米", 120, 92, 24, 2, "77%", "20%", delayed_start, delayed_finish, "张工", "机电一队", "稳定严重滞后样本"],
        ["S-002", "A座", "2层", "机电", "电气", "桥架安装", "米", 120, 70, 72, 12, "58%", "60%", normal_start, (today + timedelta(days=20)).isoformat(), "李工", "机电一队", "正常"],
        ["S-003", "B1", "3层", "消防", "喷淋系统", "喷淋主管安装", "米", 90, 78, 18, 1, "87%", "20%", delayed_start, overdue_finish, "王工", "消防一队", "稳定严重滞后样本"],
        ["S-004", "B座", "2层", "消防", "消火栓系统", "消火栓箱安装", "套", 30, 20, 24, 4, "67%", "80%", normal_start, (today + timedelta(days=25)).isoformat(), "赵工", "消防一队", "超前"],
        ["S-005", "A1", "4层", "智能化", "综合布线", "智能化桥架敷设", "米", 110, 74, 16, 1, "67%", "15%", delayed_start, delayed_finish, "钱工", "智能化一队", "稳定明显滞后样本"],
        ["S-006", "C座", "2层", "智能化", "安防系统", "摄像机安装", "台", 40, 20, 18, 3, "50%", "45%", normal_start, (today + timedelta(days=30)).isoformat(), "孙工", "智能化一队", "轻微滞后"],
    ]


def discipline_rows(discipline: str, building: str, system_name: str) -> list[list[object]]:
    today = date.today()
    delayed_start = (today - timedelta(days=14)).isoformat()
    task_prefix = {
        "机电": "管线安装",
        "消防": "消防末端安装",
        "智能化": "弱电终端安装",
    }[discipline]
    return [
        [f"{discipline}-001", building, "1层", discipline, system_name, f"{task_prefix} 1层", "米", 80, 40, 44, 8, "50%", "55%", today.isoformat(), (today + timedelta(days=12)).isoformat(), "自动验收", f"{discipline}单位", "正常"],
        [f"{discipline}-002", building, "2层", discipline, system_name, f"{task_prefix} 2层", "米", 100, 78, 28, 2, "78%", "28%", delayed_start, (today + timedelta(days=4)).isoformat(), "自动验收", f"{discipline}单位", "稳定滞后样本"],
    ]


def abnormal_rows() -> list[list[object]]:
    today = date.today()
    return [
        ["E-001", "D座", "1层", "机电", "通风", "风管安装", "米", 100, 60, 50, 10, "60%", "50%", today.isoformat(), (today + timedelta(days=15)).isoformat(), "异常", "测试单位", "正常行"],
        ["E-002", "D座", "2层", "机电", "通风", "风管支架安装", "米", -20, 10, 5, 1, "50%", "25%", today.isoformat(), (today + timedelta(days=15)).isoformat(), "异常", "测试单位", "负数工程量，应为 error"],
        ["E-003", "D座", "3层", "机电", "通风", "风口安装", "个", 20, 10, 30, 5, "50%", "150%", today.isoformat(), "bad-date", "异常", "测试单位", "warning 行"],
    ]


def write_merged_header_sheet(workbook: Workbook) -> None:
    sheet = workbook.create_sheet("06_合并表头样例")
    sheet.append(["WBS编码", "楼栋", "楼层", "专业", "系统", "工作内容", "单位", "工程量", "", "", "", "完成率", "", "日期", "", "责任", "", "备注"])
    sheet.append(["", "", "", "", "", "", "", "总工程量", "计划完成量", "累计完成量", "本期完成量", "计划完成率", "实际完成率", "计划开始", "计划完成", "责任人", "施工单位", ""])
    sheet.merge_cells("H1:K1")
    sheet.merge_cells("L1:M1")
    sheet.merge_cells("N1:O1")
    sheet.merge_cells("P1:Q1")
    for row in standard_rows()[:2]:
        sheet.append(row)
    style_header(sheet, rows=2)


def write_non_progress_field_sheet(workbook: Workbook) -> None:
    sheet = workbook.create_sheet("07_非进度Sheet_字段检查")
    sheet.append(["字段名称", "字段类型", "是否必填", "说明"])
    sheet.append(["项目名称", "文本", "是", "用于说明，不是进度明细"])
    sheet.append(["施工单位", "文本", "否", "用于说明，不是进度明细"])
    style_header(sheet)


def write_non_progress_issue_sheet(workbook: Workbook) -> None:
    sheet = workbook.create_sheet("08_非进度Sheet_问题记录")
    sheet.append(["问题编号", "问题描述", "责任人", "处理状态"])
    sheet.append(["Q-001", "现场材料待确认", "张工", "处理中"])
    sheet.append(["Q-002", "图纸会审待关闭", "李工", "未开始"])
    style_header(sheet)


def write_helper_sheet(workbook: Workbook) -> None:
    sheet = workbook.create_sheet("09_计划基线测试")
    sheet.append(["工作内容", "备注"])
    sheet.append(["合计", "辅助说明行，应被跳过"])
    sheet.append(["小计", "辅助说明行，应被跳过"])
    sheet.append(["总计", "辅助说明行，应被跳过"])
    style_header(sheet)


def style_header(sheet, rows: int = 1) -> None:
    fill = PatternFill("solid", fgColor="D9EAF7")
    for row in sheet.iter_rows(min_row=1, max_row=rows):
        for cell in row:
            cell.font = Font(bold=True)
            cell.fill = fill


def autosize(sheet) -> None:
    for column_cells in sheet.columns:
        max_length = 8
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, min(len(value) + 2, 28))
        sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = max_length


def freeze_header(sheet) -> None:
    sheet.freeze_panes = "A2"
    if sheet.title == "06_合并表头样例":
        sheet.freeze_panes = "A3"


if __name__ == "__main__":
    main()
