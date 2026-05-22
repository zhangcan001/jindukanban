import json
from datetime import time
from pathlib import Path
import tempfile
from io import BytesIO

from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook

from app.database import SessionLocal
from app.main import app
from app.models.baseline_plan import BaselinePlan
from app.models.import_batch import ImportBatch
from app.models.mapping_template import MappingTemplate
from app.models.progress_item import ProgressItem


SAMPLE_DIR = Path(__file__).resolve().parents[2] / "samples"


FIELD_MAPPINGS_A = [
    {"excel_column_name": "WBS编码", "system_field_name": "wbs_code", "field_type": "text", "is_dimension": True},
    {"excel_column_name": "楼栋", "system_field_name": "building", "field_type": "text", "is_dimension": True},
    {"excel_column_name": "楼层", "system_field_name": "floor", "field_type": "text", "is_dimension": True},
    {"excel_column_name": "专业", "system_field_name": "discipline", "field_type": "text", "is_dimension": True},
    {"excel_column_name": "工作内容", "system_field_name": "task_name", "field_type": "text", "is_dimension": True},
    {"excel_column_name": "单位", "system_field_name": "unit", "field_type": "text", "is_dimension": True},
    {"excel_column_name": "总工程量", "system_field_name": "total_quantity", "field_type": "number", "is_metric": True},
    {"excel_column_name": "计划完成量", "system_field_name": "planned_quantity", "field_type": "number", "is_metric": True},
    {"excel_column_name": "累计完成量", "system_field_name": "cumulative_quantity", "field_type": "number", "is_metric": True},
    {"excel_column_name": "本周完成量", "system_field_name": "period_quantity", "field_type": "number", "is_metric": True},
    {"excel_column_name": "计划完成率", "system_field_name": "planned_percent", "field_type": "percent", "is_metric": True},
    {"excel_column_name": "实际完成率", "system_field_name": "actual_percent", "field_type": "percent", "is_metric": True},
    {"excel_column_name": "备注", "system_field_name": "remark", "field_type": "text", "save_to_extra": True},
]


FIELD_MAPPINGS_WEIGHT = [
    {"excel_column_name": "工作内容", "system_field_name": "task_name", "field_type": "text", "is_dimension": True},
    {"excel_column_name": "实际完成率", "system_field_name": "actual_percent", "field_type": "percent", "is_metric": True},
    {"excel_column_name": "计划完成率", "system_field_name": "planned_percent", "field_type": "percent", "is_metric": True},
    {"excel_column_name": "权重", "system_field_name": "weight", "field_type": "number", "is_metric": True, "save_to_extra": True},
]


def test_validate_missing_batch_returns_clear_error() -> None:
    with TestClient(app) as client:
        response = client.post("/api/imports/999999/validate", json={"field_mappings": FIELD_MAPPINGS_A})

    assert response.status_code == 404
    assert response.json()["detail"] == "Import batch not found"


def test_t030_weight_values_are_normalized_on_import(tmp_path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["工作内容", "实际完成率", "计划完成率", "权重"])
    sheet.append(["百分比格式", "50%", "60%", 0.25])
    sheet["D2"].number_format = "0%"
    sheet.append(["文本百分比", "60%", "70%", "25%"])
    sheet.append(["普通数字", "70%", "80%", 25])
    sheet.append(["小数", "80%", "90%", 0.25])
    file_path = tmp_path / "weight.xlsx"
    workbook.save(file_path)

    with TestClient(app) as client:
        project_id = client.post("/api/projects", json={"name": "T030 权重导入"}).json()["id"]
        with file_path.open("rb") as file:
            upload_response = client.post(
                f"/api/projects/{project_id}/imports/upload",
                data={"data_date": "2026-05-18"},
                files={"file": ("weight.xlsx", file, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            )
        batch_id = upload_response.json()["batch"]["id"]
        sheet_name = upload_response.json()["sheets"][0]
        client.post(f"/api/imports/{batch_id}/parse", json={"sheet_name": sheet_name, "header_row_index": 1, "data_start_row_index": 2})
        confirm_response = client.post(f"/api/imports/{batch_id}/confirm", json={"import_strategy": "new_batch", "field_mappings": FIELD_MAPPINGS_WEIGHT})
        assert confirm_response.status_code == 200

    db = SessionLocal()
    try:
        weights = [item.weight for item in db.query(ProgressItem).filter(ProgressItem.batch_id == batch_id).order_by(ProgressItem.id).all()]
    finally:
        db.close()

    assert weights == [0.25, 0.25, 0.25, 0.25]


def test_validate_requires_selected_sheet() -> None:
    with TestClient(app) as client:
        project_id = client.post("/api/projects", json={"name": "未解析Sheet校验"}).json()["id"]
        sample_path = SAMPLE_DIR / "sample_progress_a.csv"
        with sample_path.open("rb") as file:
            upload_response = client.post(
                f"/api/projects/{project_id}/imports/upload",
                files={"file": ("sample_progress_a.csv", file, "text/csv")},
            )
        batch_id = upload_response.json()["batch"]["id"]

        response = client.post(f"/api/imports/{batch_id}/validate", json={"field_mappings": FIELD_MAPPINGS_A})

    assert response.status_code == 400
    assert response.json()["detail"]["code"] == "SHEET_NOT_SELECTED"
    assert response.json()["detail"]["message"] == "请先选择并解析要导入的 Sheet"


def test_validate_rejects_empty_field_mappings() -> None:
    with TestClient(app) as client:
        project_id = client.post("/api/projects", json={"name": "空映射校验"}).json()["id"]
        batch_id = _upload_and_parse_sample(client, project_id, "2026-05-13")

        response = client.post(f"/api/imports/{batch_id}/validate", json={"field_mappings": []})

    assert response.status_code == 400
    assert response.json()["detail"]["code"] == "FIELD_MAPPINGS_EMPTY"
    assert response.json()["detail"]["message"] == "导入校验失败：字段映射不能为空。"


def test_validate_normal_request_returns_validation_result() -> None:
    with TestClient(app) as client:
        project_id = client.post("/api/projects", json={"name": "正常校验返回"}).json()["id"]
        batch_id = _upload_and_parse_sample(client, project_id, "2026-05-13")

        response = client.post(f"/api/imports/{batch_id}/validate", json={"field_mappings": FIELD_MAPPINGS_A})

    assert response.status_code == 200
    assert response.json()["valid"] is True
    assert "data_quality" in response.json()
    assert "issues" in response.json()
    assert "abnormal_preview" in response.json()


def test_validate_time_cell_in_date_field_returns_issue_instead_of_500(tmp_path: Path) -> None:
    sample_path = tmp_path / "date-time-cell.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "日期异常"
    sheet.append(["工作内容", "计划完成"])
    sheet.append(["桥架安装", time(8, 30)])
    workbook.save(sample_path)
    mappings = [
        {"excel_column_name": "工作内容", "system_field_name": "task_name", "field_type": "text"},
        {"excel_column_name": "计划完成", "system_field_name": "planned_finish_date", "field_type": "date"},
    ]

    with TestClient(app) as client:
        project_id = client.post("/api/projects", json={"name": "日期时间异常校验"}).json()["id"]
        with sample_path.open("rb") as file:
            upload_response = client.post(
                f"/api/projects/{project_id}/imports/upload",
                files={"file": (sample_path.name, file, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            )
        batch_id = upload_response.json()["batch"]["id"]
        parse_response = client.post(
            f"/api/imports/{batch_id}/parse",
            json={"sheet_name": "日期异常", "header_row_index": 1, "data_start_row_index": 2},
        )
        response = client.post(f"/api/imports/{batch_id}/validate", json={"field_mappings": mappings})

    assert parse_response.status_code == 200
    assert parse_response.json()["header_recommendation"]["header_row_index"] == 1
    assert parse_response.json()["header_recommendation"]["data_start_row_index"] == 2
    assert response.status_code == 200
    assert any(issue["code"] == "INVALID_DATE" for issue in response.json()["issues"])
    assert response.json()["abnormal_preview"][0]["type"] == "日期异常"


def test_abnormal_preview_groups_common_issue_types(tmp_path: Path) -> None:
    sample_path = tmp_path / "abnormal-preview.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "异常预览"
    sheet.append(["工作内容", "单位", "总工程量", "实际完成率", "计划完成"])
    sheet.append(["桥架安装", "米", -1, "120%", "bad-date"])
    sheet.append(["合计", "米", 10, "50%", "2026-05-18"])
    workbook.save(sample_path)
    mappings = [
        {"excel_column_name": "工作内容", "system_field_name": "task_name", "field_type": "text"},
        {"excel_column_name": "单位", "system_field_name": "unit", "field_type": "text"},
        {"excel_column_name": "总工程量", "system_field_name": "total_quantity", "field_type": "number"},
        {"excel_column_name": "实际完成率", "system_field_name": "actual_percent", "field_type": "percent"},
        {"excel_column_name": "计划完成", "system_field_name": "planned_finish_date", "field_type": "date"},
    ]

    with TestClient(app) as client:
        project_id = client.post("/api/projects", json={"name": "异常预览分组"}).json()["id"]
        with sample_path.open("rb") as file:
            upload_response = client.post(
                f"/api/projects/{project_id}/imports/upload",
                files={"file": (sample_path.name, file, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            )
        batch_id = upload_response.json()["batch"]["id"]
        client.post(f"/api/imports/{batch_id}/parse", json={"sheet_name": "异常预览", "header_row_index": 1, "data_start_row_index": 2})
        response = client.post(f"/api/imports/{batch_id}/validate", json={"field_mappings": mappings})

    assert response.status_code == 200
    groups = {group["type"]: group for group in response.json()["abnormal_preview"]}
    assert groups["日期异常"]["examples"][0]["raw_value"] == "bad-date"
    assert groups["负数工程量"]["level"] == "error"
    assert groups["完成率超范围"]["count"] == 1
    assert groups["汇总行跳过"]["count"] == 1


def test_confirm_blocks_row_level_errors_and_publish_remains_blocked(tmp_path: Path) -> None:
    sample_path = tmp_path / "row-error.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "05_异常数据校验"
    sheet.append(["工作内容", "单位", "总工程量", "累计完成量", "实际完成率"])
    sheet.append(["桥架安装", "米", 100, 50, "50%"])
    sheet.append(["风管安装", "米", -20, 5, "25%"])
    workbook.save(sample_path)
    mappings = [
        {"excel_column_name": "工作内容", "system_field_name": "task_name", "field_type": "text"},
        {"excel_column_name": "单位", "system_field_name": "unit", "field_type": "text"},
        {"excel_column_name": "总工程量", "system_field_name": "total_quantity", "field_type": "number"},
        {"excel_column_name": "累计完成量", "system_field_name": "cumulative_quantity", "field_type": "number"},
        {"excel_column_name": "实际完成率", "system_field_name": "actual_percent", "field_type": "percent"},
    ]

    with TestClient(app) as client:
        project_id = client.post("/api/projects", json={"name": "error发布拦截"}).json()["id"]
        with sample_path.open("rb") as file:
            upload_response = client.post(
                f"/api/projects/{project_id}/imports/upload",
                files={"file": (sample_path.name, file, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            )
        batch_id = upload_response.json()["batch"]["id"]
        client.post(f"/api/imports/{batch_id}/parse", json={"sheet_name": "05_异常数据校验", "header_row_index": 1, "data_start_row_index": 2})
        validate_response = client.post(f"/api/imports/{batch_id}/validate", json={"field_mappings": mappings})
        confirm_response = client.post(f"/api/imports/{batch_id}/confirm", json={"import_strategy": "new_batch", "field_mappings": mappings})
        publish_response = client.post(f"/api/imports/{batch_id}/publish")

    assert validate_response.status_code == 200
    assert validate_response.json()["valid"] is False
    assert confirm_response.status_code == 200
    assert confirm_response.json()["valid"] is False
    assert confirm_response.json()["imported_count"] == 0
    assert confirm_response.json()["error_count"] == 1
    assert publish_response.status_code == 400
    db = SessionLocal()
    try:
        assert db.query(ProgressItem).filter(ProgressItem.batch_id == batch_id).count() == 0
    finally:
        db.close()


def test_confirm_and_publish_block_zero_importable_rows(tmp_path: Path) -> None:
    sample_path = tmp_path / "helper-sheet.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "09_计划基线测试"
    sheet.append(["工作内容", "备注"])
    sheet.append(["合计", "辅助说明"])
    sheet.append(["小计", "辅助说明"])
    sheet.append(["总计", "辅助说明"])
    workbook.save(sample_path)
    mappings = [
        {"excel_column_name": "工作内容", "system_field_name": "task_name", "field_type": "text"},
        {"excel_column_name": "备注", "system_field_name": "remark", "field_type": "text"},
    ]

    with TestClient(app) as client:
        project_id = client.post("/api/projects", json={"name": "辅助Sheet拦截"}).json()["id"]
        with sample_path.open("rb") as file:
            upload_response = client.post(
                f"/api/projects/{project_id}/imports/upload",
                files={"file": (sample_path.name, file, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            )
        batch_id = upload_response.json()["batch"]["id"]
        client.post(f"/api/imports/{batch_id}/parse", json={"sheet_name": "09_计划基线测试", "header_row_index": 1, "data_start_row_index": 2})
        confirm_response = client.post(f"/api/imports/{batch_id}/confirm", json={"import_strategy": "new_batch", "field_mappings": mappings})
        publish_response = client.post(f"/api/imports/{batch_id}/publish")

    assert confirm_response.status_code == 200
    assert confirm_response.json()["valid"] is False
    assert confirm_response.json()["imported_count"] == 0
    assert any(issue["code"] == "NO_IMPORTABLE_ROWS" for issue in confirm_response.json()["issues"])
    assert publish_response.status_code == 400
    assert "未生成有效进度数据" in str(publish_response.json()["detail"])


def test_merged_header_parse_validate_confirm_persists_auto_multi_header(tmp_path: Path) -> None:
    sample_path = tmp_path / "merged-full-flow.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "06_合并表头样例"
    sheet.append(["楼栋", "楼层", "专业", "系统", "施工项", "单位", "工程量", "", "", "", "完成率", ""])
    sheet.append(["", "", "", "", "", "", "总工程量", "计划完成量", "累计完成量", "本期完成量", "计划完成率", "实际完成率"])
    sheet.merge_cells("G1:J1")
    sheet.merge_cells("K1:L1")
    sheet.append(["1号楼", "1层", "机电", "给排水", "给水管安装", "米", 100, 60, 50, 10, "60%", "50%"])
    sheet.append(["1号楼", "2层", "机电", "电气", "桥架安装", "米", 200, 120, 90, 20, "60%", "45%"])
    workbook.save(sample_path)

    with TestClient(app) as client:
        project_id = client.post("/api/projects", json={"name": "合并表头全链路"}).json()["id"]
        with sample_path.open("rb") as file:
            upload_response = client.post(
                f"/api/projects/{project_id}/imports/upload",
                files={"file": (sample_path.name, file, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            )
        batch_id = upload_response.json()["batch"]["id"]
        parse_response = client.post(
            f"/api/imports/{batch_id}/parse",
            json={"sheet_name": "06_合并表头样例", "header_row_index": 1, "data_start_row_index": None, "multi_header": False},
        )
        columns = parse_response.json()["columns"]
        mappings = [
            {
                "excel_column_name": column["name"],
                "system_field_name": column["recommended_field"],
                "field_type": column["field_type"],
                "is_dimension": column["is_dimension"],
                "is_metric": column["is_metric"],
                "save_to_extra": column["save_to_extra"],
            }
            for column in columns
            if column["recommended_field"]
        ]
        validate_response = client.post(f"/api/imports/{batch_id}/validate", json={"field_mappings": mappings})
        confirm_response = client.post(f"/api/imports/{batch_id}/confirm", json={"import_strategy": "new_batch", "field_mappings": mappings})

    assert parse_response.status_code == 200
    assert parse_response.json()["batch"]["multi_header"] is True
    assert parse_response.json()["batch"]["header_end_row_index"] == 2
    assert validate_response.status_code == 200
    assert validate_response.json()["valid"] is True
    assert confirm_response.status_code == 200
    assert confirm_response.json()["valid"] is True
    assert confirm_response.json()["imported_count"] == 2


def test_full_csv_import_publish_analytics_and_report_flow() -> None:
    with TestClient(app) as client:
        project_response = client.post("/api/projects", json={"name": "API导入链路测试", "project_type": "测试"})
        assert project_response.status_code == 201
        project_id = project_response.json()["id"]

        sample_path = SAMPLE_DIR / "sample_progress_a.csv"
        with sample_path.open("rb") as file:
            upload_response = client.post(
                f"/api/projects/{project_id}/imports/upload",
                data={"data_date": "2026-05-13"},
                files={"file": ("sample_progress_a.csv", file, "text/csv")},
            )
        assert upload_response.status_code == 200
        batch_id = upload_response.json()["batch"]["id"]
        assert upload_response.json()["batch"]["data_date"] == "2026-05-13"
        assert upload_response.json()["sheets"] == ["CSV"]

        parse_response = client.post(
            f"/api/imports/{batch_id}/parse",
            json={"sheet_name": "CSV", "header_row_index": 1, "data_start_row_index": 2},
        )
        assert parse_response.status_code == 200
        assert parse_response.json()["batch"]["row_count"] == 5

        mapping_response = client.post(
            f"/api/imports/{batch_id}/mapping/validate",
            json={"field_mappings": FIELD_MAPPINGS_A},
        )
        assert mapping_response.status_code == 200
        assert mapping_response.json()["valid"] is True

        validate_response = client.post(
            f"/api/imports/{batch_id}/validate",
            json={"field_mappings": FIELD_MAPPINGS_A},
        )
        assert validate_response.status_code == 200
        assert validate_response.json()["valid"] is True
        assert validate_response.json()["data_quality"]["data_quality_score"] > 70

        confirm_response = client.post(
            f"/api/imports/{batch_id}/confirm",
            json={
                "save_as_template": True,
                "template_name": "pytest标准模板",
                "data_date": "2026-05-14",
                "import_strategy": "new_batch",
                "field_mappings": FIELD_MAPPINGS_A,
            },
        )
        assert confirm_response.status_code == 200
        assert confirm_response.json()["status"] == "imported"
        assert confirm_response.json()["imported_count"] == 5
        db = SessionLocal()
        try:
            assert db.get(ImportBatch, batch_id).data_date.isoformat() == "2026-05-14"
        finally:
            db.close()

        publish_response = client.post(f"/api/imports/{batch_id}/publish")
        assert publish_response.status_code == 200
        assert publish_response.json()["status"] == "published"

        overview_response = client.get(f"/api/projects/{project_id}/analytics/overview", params={"batch_id": batch_id})
        assert overview_response.status_code == 200
        overview = overview_response.json()
        assert overview["item_count"] == 5
        assert overview["actual_percent"] is not None

        report_response = client.get(f"/api/projects/{project_id}/reports/overview", params={"batch_id": batch_id})
        assert report_response.status_code == 200
        assert report_response.content[:2] == b"PK"
        assert "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet" in report_response.headers["content-type"]


def test_replace_same_date_deactivates_old_published_batch_and_trend_sorts_active_batches() -> None:
    with TestClient(app) as client:
        project_response = client.post("/api/projects", json={"name": "同日期替换测试"})
        project_id = project_response.json()["id"]

        first_batch_id = _import_and_publish_sample(client, project_id, "2026-05-06", "new_batch")
        inactive_batch_id = _import_and_publish_sample(client, project_id, "2026-05-13", "new_batch")
        replacement_batch_id = _import_and_publish_sample(client, project_id, "2026-05-13", "replace_same_date")

        trend_response = client.get(f"/api/projects/{project_id}/analytics/trend")

    db = SessionLocal()
    try:
        assert db.get(ImportBatch, inactive_batch_id).is_active is False
        assert db.get(ImportBatch, replacement_batch_id).is_active is True
    finally:
        db.close()

    assert trend_response.status_code == 200
    rows = trend_response.json()["rows"]
    assert [row["batch_id"] for row in rows] == [first_batch_id, replacement_batch_id]
    assert [row["data_date"] for row in rows] == ["2026-05-06", "2026-05-13"]


def test_freeze_batch_blocks_replace_same_date_and_can_unfreeze() -> None:
    with TestClient(app) as client:
        project_response = client.post("/api/projects", json={"name": "冻结批次测试"})
        project_id = project_response.json()["id"]

        frozen_batch_id = _import_and_publish_sample(client, project_id, "2026-05-14", "new_batch")
        freeze_response = client.post(f"/api/imports/{frozen_batch_id}/freeze", json={"freeze_remark": "月度封版"})
        replacement_id = _upload_and_parse_sample(client, project_id, "2026-05-14")
        blocked = client.post(
            f"/api/imports/{replacement_id}/confirm",
            json={
                "data_date": "2026-05-14",
                "import_strategy": "replace_same_date",
                "field_mappings": FIELD_MAPPINGS_A,
            },
        )
        frozen_overwrite = client.post(
            f"/api/imports/{frozen_batch_id}/confirm",
            json={
                "data_date": "2026-05-14",
                "import_strategy": "overwrite_current",
                "field_mappings": FIELD_MAPPINGS_A,
            },
        )
        progress_items = client.get(f"/api/projects/{project_id}/progress-items", params={"batch_id": frozen_batch_id}).json()["items"]
        manual_edit_blocked = client.put(
            f"/api/progress-items/{progress_items[0]['id']}",
            json={"actual_percent": 80, "reason": "冻结保护测试"},
        )
        report_allowed = client.get(f"/api/projects/{project_id}/reports/weekly-word", params={"batch_id": frozen_batch_id})
        unfreeze_response = client.post(f"/api/imports/{frozen_batch_id}/unfreeze")
        allowed = client.post(
            f"/api/imports/{replacement_id}/confirm",
            json={
                "data_date": "2026-05-14",
                "import_strategy": "replace_same_date",
                "field_mappings": FIELD_MAPPINGS_A,
            },
        )

    assert freeze_response.status_code == 200
    assert freeze_response.json()["is_frozen"] is True
    assert blocked.status_code == 400
    assert "冻结批次" in str(blocked.json()["detail"])
    assert frozen_overwrite.status_code == 400
    assert manual_edit_blocked.status_code == 400
    assert report_allowed.status_code == 200
    assert unfreeze_response.status_code == 200
    assert unfreeze_response.json()["is_frozen"] is False
    assert allowed.status_code == 200

    db = SessionLocal()
    try:
        assert db.get(ImportBatch, frozen_batch_id).is_active is False
    finally:
        db.close()


def test_extra_fields_respects_save_to_extra_for_mapped_unmapped_and_ignored_columns() -> None:
    with TestClient(app) as client:
        project_response = client.post("/api/projects", json={"name": "扩展字段测试"})
        project_id = project_response.json()["id"]
        batch_id = _upload_and_parse_sample(client, project_id, "2026-05-13")
        mappings = [
            {**mapping, "save_to_extra": False}
            for mapping in FIELD_MAPPINGS_A
            if mapping["excel_column_name"] != "备注"
        ]
        mappings.extend(
            [
                {"excel_column_name": "备注", "system_field_name": "remark", "field_type": "text", "save_to_extra": True},
                {"excel_column_name": "楼栋", "system_field_name": "building", "field_type": "text", "save_to_extra": True},
                {"excel_column_name": "专业", "system_field_name": "discipline", "field_type": "text", "save_to_extra": False},
            ]
        )
        # Use one mapping per standard field; override duplicates by rebuilding explicitly.
        mappings = [
            {"excel_column_name": "WBS编码", "system_field_name": "wbs_code", "field_type": "text", "save_to_extra": False},
            {"excel_column_name": "楼栋", "system_field_name": "building", "field_type": "text", "save_to_extra": True},
            {"excel_column_name": "楼层", "system_field_name": "floor", "field_type": "text", "save_to_extra": False},
            {"excel_column_name": "专业", "system_field_name": "discipline", "field_type": "text", "save_to_extra": False},
            {"excel_column_name": "工作内容", "system_field_name": "task_name", "field_type": "text", "save_to_extra": False},
            {"excel_column_name": "单位", "system_field_name": "unit", "field_type": "text", "save_to_extra": False},
            {"excel_column_name": "总工程量", "system_field_name": "total_quantity", "field_type": "number", "save_to_extra": False},
            {"excel_column_name": "累计完成量", "system_field_name": "cumulative_quantity", "field_type": "number", "save_to_extra": False},
            {"excel_column_name": "计划完成率", "system_field_name": "planned_percent", "field_type": "percent", "save_to_extra": False},
            {"excel_column_name": "实际完成率", "system_field_name": "actual_percent", "field_type": "percent", "save_to_extra": False},
            {"excel_column_name": "备注", "system_field_name": "remark", "field_type": "text", "save_to_extra": True},
            {"excel_column_name": "计划完成量", "system_field_name": None, "field_type": "number", "save_to_extra": True},
            {"excel_column_name": "本周完成量", "system_field_name": None, "field_type": "number", "save_to_extra": False},
        ]

        confirm_response = client.post(
            f"/api/imports/{batch_id}/confirm",
            json={"data_date": "2026-05-13", "import_strategy": "new_batch", "field_mappings": mappings},
        )

    assert confirm_response.status_code == 200
    db = SessionLocal()
    try:
        item = db.query(ProgressItem).filter(ProgressItem.batch_id == batch_id).order_by(ProgressItem.id).first()
        extra = json.loads(item.extra_fields)
    finally:
        db.close()

    assert "计划完成量" in extra
    assert "楼栋" in extra
    assert "备注" in extra
    assert "专业" not in extra
    assert "本周完成量" not in extra


def test_publish_blocks_error_batches_and_zero_import_batches(tmp_path: Path) -> None:
    workbook_path = tmp_path / "publish-block.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "05_异常数据校验"
    sheet.append(["工作内容", "单位", "总工程量", "累计完成量", "实际完成率"])
    sheet.append(["桥架安装", "米", 100, 50, "50%"])
    sheet.append(["风管安装", "米", -20, 5, "25%"])
    workbook.save(workbook_path)
    error_mappings = [
        {"excel_column_name": "工作内容", "system_field_name": "task_name", "field_type": "text"},
        {"excel_column_name": "单位", "system_field_name": "unit", "field_type": "text"},
        {"excel_column_name": "总工程量", "system_field_name": "total_quantity", "field_type": "number"},
        {"excel_column_name": "累计完成量", "system_field_name": "cumulative_quantity", "field_type": "number"},
        {"excel_column_name": "实际完成率", "system_field_name": "actual_percent", "field_type": "percent"},
    ]
    helper_path = tmp_path / "helper-sheet.xlsx"
    helper_book = Workbook()
    helper_sheet = helper_book.active
    helper_sheet.title = "09_计划基线测试"
    helper_sheet.append(["工作内容", "备注"])
    helper_sheet.append(["合计", "辅助说明"])
    helper_sheet.append(["小计", "辅助说明"])
    helper_sheet.append(["总计", "辅助说明"])
    helper_book.save(helper_path)
    helper_mappings = [
        {"excel_column_name": "工作内容", "system_field_name": "task_name", "field_type": "text"},
        {"excel_column_name": "备注", "system_field_name": "remark", "field_type": "text"},
    ]

    with TestClient(app) as client:
        project_id = client.post("/api/projects", json={"name": "发布拦截组合"}).json()["id"]
        with workbook_path.open("rb") as file:
            error_upload = client.post(
                f"/api/projects/{project_id}/imports/upload",
                files={"file": (workbook_path.name, file, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            )
        error_batch_id = error_upload.json()["batch"]["id"]
        client.post(f"/api/imports/{error_batch_id}/parse", json={"sheet_name": "05_异常数据校验", "header_row_index": 1, "data_start_row_index": 2})
        client.post(f"/api/imports/{error_batch_id}/validate", json={"field_mappings": error_mappings})
        error_publish = client.post(f"/api/imports/{error_batch_id}/publish")

        with helper_path.open("rb") as file:
            helper_upload = client.post(
                f"/api/projects/{project_id}/imports/upload",
                files={"file": (helper_path.name, file, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            )
        helper_batch_id = helper_upload.json()["batch"]["id"]
        client.post(f"/api/imports/{helper_batch_id}/parse", json={"sheet_name": "09_计划基线测试", "header_row_index": 1, "data_start_row_index": 2})
        helper_confirm = client.post(f"/api/imports/{helper_batch_id}/confirm", json={"import_strategy": "new_batch", "field_mappings": helper_mappings})
        helper_publish = client.post(f"/api/imports/{helper_batch_id}/publish")

    assert error_publish.status_code == 400
    assert helper_confirm.status_code == 200
    assert helper_confirm.json()["imported_count"] == 0
    assert helper_publish.status_code == 400


def test_abnormal_sample_summary_row_is_skipped_on_confirm() -> None:
    mappings = [
        {"excel_column_name": "楼栋", "system_field_name": "building", "field_type": "text"},
        {"excel_column_name": "楼层", "system_field_name": "floor", "field_type": "text"},
        {"excel_column_name": "专业", "system_field_name": "discipline", "field_type": "text"},
        {"excel_column_name": "工作内容", "system_field_name": "task_name", "field_type": "text"},
        {"excel_column_name": "单位", "system_field_name": "unit", "field_type": "text"},
        {"excel_column_name": "总工程量", "system_field_name": "total_quantity", "field_type": "number"},
        {"excel_column_name": "累计完成量", "system_field_name": "cumulative_quantity", "field_type": "number"},
        {"excel_column_name": "计划完成率", "system_field_name": "planned_percent", "field_type": "percent"},
        {"excel_column_name": "实际完成率", "system_field_name": "actual_percent", "field_type": "percent"},
        {"excel_column_name": "计划开始", "system_field_name": "planned_start_date", "field_type": "date"},
        {"excel_column_name": "计划完成", "system_field_name": "planned_finish_date", "field_type": "date"},
    ]

    with TestClient(app) as client:
        project_response = client.post("/api/projects", json={"name": "异常合计行跳过测试"})
        project_id = project_response.json()["id"]
        sample_path = SAMPLE_DIR / "sample_progress_abnormal.csv"
        with sample_path.open("rb") as file:
            upload_response = client.post(
                f"/api/projects/{project_id}/imports/upload",
                data={"data_date": "2026-05-27"},
                files={"file": ("sample_progress_abnormal.csv", file, "text/csv")},
            )
        batch_id = upload_response.json()["batch"]["id"]
        parse_response = client.post(
            f"/api/imports/{batch_id}/parse",
            json={"sheet_name": "CSV", "data_date": "2026-05-27", "header_row_index": 1, "data_start_row_index": 2},
        )
        assert parse_response.status_code == 200

        validate_response = client.post(f"/api/imports/{batch_id}/validate", json={"field_mappings": mappings})

    assert validate_response.status_code == 200
    codes = {issue["code"] for issue in validate_response.json()["issues"]}
    assert "SUMMARY_ROW_SKIPPED" in codes
    assert "task_name_empty" not in codes
    db = SessionLocal()
    try:
        batch = db.get(ImportBatch, batch_id)
        assert batch.skipped_count == 2
    finally:
        db.close()


def test_summary_row_is_saved_as_raw_but_not_imported_as_task_or_item() -> None:
    sample_path = Path(tempfile.gettempdir()) / "summary_skip_importable.csv"
    sample_path.write_text(
        "工作内容,专业,单位,总工程量,累计完成量,实际完成率\n"
        "配电箱安装,电气,台,10,5,50%\n"
        "合计,,,,10,50%\n"
        "合计箱安装,电气,台,8,4,50%\n",
        encoding="utf-8-sig",
    )
    mappings = [
        {"excel_column_name": "工作内容", "system_field_name": "task_name", "field_type": "text"},
        {"excel_column_name": "专业", "system_field_name": "discipline", "field_type": "text"},
        {"excel_column_name": "单位", "system_field_name": "unit", "field_type": "text"},
        {"excel_column_name": "总工程量", "system_field_name": "total_quantity", "field_type": "number"},
        {"excel_column_name": "累计完成量", "system_field_name": "cumulative_quantity", "field_type": "number"},
        {"excel_column_name": "实际完成率", "system_field_name": "actual_percent", "field_type": "percent"},
    ]

    with TestClient(app) as client:
        project_response = client.post("/api/projects", json={"name": "可导入合计行跳过测试"})
        project_id = project_response.json()["id"]
        with sample_path.open("rb") as file:
            upload_response = client.post(
                f"/api/projects/{project_id}/imports/upload",
                data={"data_date": "2026-05-28"},
                files={"file": ("summary_skip_importable.csv", file, "text/csv")},
            )
        batch_id = upload_response.json()["batch"]["id"]
        client.post(
            f"/api/imports/{batch_id}/parse",
            json={"sheet_name": "CSV", "data_date": "2026-05-28", "header_row_index": 1, "data_start_row_index": 2},
        )
        validate_response = client.post(f"/api/imports/{batch_id}/validate", json={"field_mappings": mappings})
        confirm_response = client.post(
            f"/api/imports/{batch_id}/confirm",
            json={"data_date": "2026-05-28", "import_strategy": "new_batch", "field_mappings": mappings},
        )

    assert validate_response.status_code == 200
    assert confirm_response.status_code == 200
    assert confirm_response.json()["imported_count"] == 2
    assert confirm_response.json()["skipped_count"] == 1
    db = SessionLocal()
    try:
        batch = db.get(ImportBatch, batch_id)
        assert batch.imported_count == 2
        assert batch.skipped_count == 1
        names = [row.task_name for row in db.query(ProgressItem).filter(ProgressItem.batch_id == batch_id).all()]
        assert names == ["配电箱安装", "合计箱安装"]
    finally:
        db.close()


def test_confirm_import_skips_row_level_errors_without_blocking_valid_rows(tmp_path: Path) -> None:
    sample_path = tmp_path / "abnormal_with_valid_rows.csv"
    sample_path.write_text(
        "工作内容,专业,单位,总工程量,累计完成量,实际完成率,计划完成\n"
        "桥架安装,电气,米,100,120,120%,2026-05-30\n"
        "电缆敷设,电气,米,-100,50,50%,日期错误\n"
        "风管安装,暖通,平方米,200,70,35%,\n"
        "合计,,,,390,60%,\n",
        encoding="utf-8-sig",
    )
    mappings = [
        {"excel_column_name": "工作内容", "system_field_name": "task_name", "field_type": "text"},
        {"excel_column_name": "专业", "system_field_name": "discipline", "field_type": "text"},
        {"excel_column_name": "单位", "system_field_name": "unit", "field_type": "text"},
        {"excel_column_name": "总工程量", "system_field_name": "total_quantity", "field_type": "number"},
        {"excel_column_name": "累计完成量", "system_field_name": "cumulative_quantity", "field_type": "number"},
        {"excel_column_name": "实际完成率", "system_field_name": "actual_percent", "field_type": "percent"},
        {"excel_column_name": "计划完成", "system_field_name": "planned_finish_date", "field_type": "date"},
    ]

    with TestClient(app) as client:
        project_response = client.post("/api/projects", json={"name": "行级错误跳过测试"})
        project_id = project_response.json()["id"]
        with sample_path.open("rb") as file:
            upload_response = client.post(
                f"/api/projects/{project_id}/imports/upload",
                data={"data_date": "2026-05-13"},
                files={"file": (sample_path.name, file, "text/csv")},
            )
        batch_id = upload_response.json()["batch"]["id"]
        parse_response = client.post(
            f"/api/imports/{batch_id}/parse",
            json={"sheet_name": "CSV", "header_row_index": 1, "data_start_row_index": 2},
        )
        validate_response = client.post(f"/api/imports/{batch_id}/validate", json={"field_mappings": mappings})
        confirm_response = client.post(
            f"/api/imports/{batch_id}/confirm",
            json={"data_date": "2026-05-13", "import_strategy": "new_batch", "field_mappings": mappings},
        )

    assert parse_response.status_code == 200
    assert validate_response.status_code == 200
    assert validate_response.json()["valid"] is False
    codes = {issue["code"] for issue in validate_response.json()["issues"]}
    assert {"negative_quantity", "INVALID_DATE", "percent_out_of_range", "actual_exceeds_total", "SUMMARY_ROW_SKIPPED"} <= codes
    assert confirm_response.status_code == 200
    assert confirm_response.json()["valid"] is False
    assert confirm_response.json()["imported_count"] == 0
    assert confirm_response.json()["skipped_count"] == 4
    assert confirm_response.json()["error_count"] == 1

    db = SessionLocal()
    try:
        batch = db.get(ImportBatch, batch_id)
        assert batch.status == "parsed"
        assert batch.imported_count == 0
        assert batch.skipped_count == 2
        names = [item.task_name for item in db.query(ProgressItem).filter(ProgressItem.batch_id == batch_id).order_by(ProgressItem.id)]
        assert names == []
    finally:
        db.close()


def test_confirm_import_skips_rows_without_progress_metrics(tmp_path: Path) -> None:
    sample_path = tmp_path / "no-progress-metrics.csv"
    sample_path.write_text(
        "工作内容,楼栋,楼层,专业,实际完成率,计划完成率\n"
        "桥架安装,1号楼,1层,电气,60%,80%\n"
        "只有任务名称,1号楼,2层,电气,,\n"
        ",,,,\n",
        encoding="utf-8-sig",
    )
    mappings = [
        {"excel_column_name": "工作内容", "system_field_name": "task_name", "field_type": "text"},
        {"excel_column_name": "楼栋", "system_field_name": "building", "field_type": "text"},
        {"excel_column_name": "楼层", "system_field_name": "floor", "field_type": "text"},
        {"excel_column_name": "专业", "system_field_name": "discipline", "field_type": "text"},
        {"excel_column_name": "实际完成率", "system_field_name": "actual_percent", "field_type": "percent"},
        {"excel_column_name": "计划完成率", "system_field_name": "planned_percent", "field_type": "percent"},
    ]

    with TestClient(app) as client:
        project_id = client.post("/api/projects", json={"name": "无进度指标跳过"}).json()["id"]
        with sample_path.open("rb") as file:
            upload_response = client.post(
                f"/api/projects/{project_id}/imports/upload",
                files={"file": (sample_path.name, file, "text/csv")},
            )
        batch_id = upload_response.json()["batch"]["id"]
        parse_response = client.post(
            f"/api/imports/{batch_id}/parse",
            json={"sheet_name": "CSV", "header_row_index": 1, "data_start_row_index": 2},
        )
        validate_response = client.post(f"/api/imports/{batch_id}/validate", json={"field_mappings": mappings})
        confirm_response = client.post(
            f"/api/imports/{batch_id}/confirm",
            json={"data_date": "2026-05-13", "import_strategy": "new_batch", "field_mappings": mappings},
        )

    assert parse_response.status_code == 200
    assert validate_response.status_code == 200
    assert any(issue["code"] == "NO_PROGRESS_METRICS" for issue in validate_response.json()["issues"])
    assert confirm_response.status_code == 200
    assert confirm_response.json()["imported_count"] == 1
    assert confirm_response.json()["skipped_count"] == 2

    db = SessionLocal()
    try:
        items = db.query(ProgressItem).filter(ProgressItem.batch_id == batch_id).all()
        assert len(items) == 1
        assert items[0].task_name == "桥架安装"
        assert items[0].actual_percent == 60
    finally:
        db.close()


def test_abnormal_and_summary_sheet_marks_and_skips_error_rows(tmp_path: Path) -> None:
    sample_path = tmp_path / "abnormal-summary.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "05_异常与合计行"
    sheet.append(["工作内容", "专业", "单位", "总工程量", "累计完成量", "实际完成率"])
    sheet.append(["桥架安装", "电气", "米", 100, 50, "50%"])
    sheet.append(["电缆敷设", "电气", "米", -100, 50, "50%"])
    sheet.append(["配电箱安装", "电气", "台", 10, 5, "50%"])
    sheet.append(["合计", "", "", 10, 5, "50%"])
    sheet.append(["专业小计", "", "", 10, 5, "50%"])
    sheet.append(["汇总", "", "", 10, 5, "50%"])
    sheet.append(["合计箱安装", "电气", "台", 8, 4, "50%"])
    workbook.save(sample_path)
    mappings = [
        {"excel_column_name": "工作内容", "system_field_name": "task_name", "field_type": "text"},
        {"excel_column_name": "专业", "system_field_name": "discipline", "field_type": "text"},
        {"excel_column_name": "单位", "system_field_name": "unit", "field_type": "text"},
        {"excel_column_name": "总工程量", "system_field_name": "total_quantity", "field_type": "number"},
        {"excel_column_name": "累计完成量", "system_field_name": "cumulative_quantity", "field_type": "number"},
        {"excel_column_name": "实际完成率", "system_field_name": "actual_percent", "field_type": "percent"},
    ]

    with TestClient(app) as client:
        project_response = client.post("/api/projects", json={"name": "05异常合计行回归"})
        project_id = project_response.json()["id"]
        with sample_path.open("rb") as file:
            upload_response = client.post(
                f"/api/projects/{project_id}/imports/upload",
                data={"data_date": "2026-05-13"},
                files={"file": (sample_path.name, file, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            )
        batch_id = upload_response.json()["batch"]["id"]
        parse_response = client.post(
            f"/api/imports/{batch_id}/parse",
            json={"sheet_name": "05_异常与合计行", "header_row_index": 1, "data_start_row_index": 2},
        )
        validate_response = client.post(f"/api/imports/{batch_id}/validate", json={"field_mappings": mappings})
        confirm_response = client.post(
            f"/api/imports/{batch_id}/confirm",
            json={"data_date": "2026-05-13", "import_strategy": "new_batch", "field_mappings": mappings},
        )
        publish_response = client.post(f"/api/imports/{batch_id}/publish")

    assert parse_response.status_code == 200
    assert validate_response.status_code == 200
    issues = validate_response.json()["issues"]
    normalized_rows = validate_response.json()["normalized_preview_rows"]
    issue_codes = [issue["code"] for issue in issues]
    assert "negative_quantity" in issue_codes
    assert issue_codes.count("SUMMARY_ROW_SKIPPED") == 3
    rows_by_name = {row["task_name"]: row for row in normalized_rows}
    assert rows_by_name["电缆敷设"]["__skip_import"] is True
    assert rows_by_name["合计"]["__skip_import"] is True
    assert rows_by_name["专业小计"]["__skip_import"] is True
    assert rows_by_name["汇总"]["__skip_import"] is True
    assert rows_by_name["合计箱安装"].get("__skip_import") is not True

    assert confirm_response.status_code == 200
    assert confirm_response.json()["valid"] is False
    assert confirm_response.json()["imported_count"] == 0
    assert confirm_response.json()["skipped_count"] == 7
    assert publish_response.status_code == 400

    db = SessionLocal()
    try:
        batch = db.get(ImportBatch, batch_id)
        assert batch.imported_count == 0
        assert batch.skipped_count == 4
        names = [item.task_name for item in db.query(ProgressItem).filter(ProgressItem.batch_id == batch_id).order_by(ProgressItem.id)]
    finally:
        db.close()
    assert names == []


def test_multi_sheet_upload_parse_and_confirm_uses_selected_sheet(tmp_path: Path) -> None:
    workbook_path = tmp_path / "v1.3_beta_工程进度测试表格包.xlsx"
    workbook = Workbook()
    workbook.active.title = "使用说明"
    workbook.active.append(["说明"])
    first_sheet = workbook.create_sheet("01_标准机电进度表")
    first_sheet.append(["WBS编码", "工作内容", "实际完成率"])
    first_sheet.append(["JD.FIRST", "第一个Sheet任务", "10%"])
    second_sheet = workbook.create_sheet("02_字段不规范表")
    second_sheet.append(["区域", "系统", "子项", "清单量", "累计完成", "本期完成"])
    second_sheet.append(["地下室", "消防系统", "喷淋主管安装", 300, 120, 30])
    abnormal_sheet = workbook.create_sheet("04_异常与合计测试表")
    abnormal_sheet.append(["楼栋", "楼层", "专业", "工作内容", "总工程量"])
    abnormal_sheet.append(["1号楼", "负一层", "电气", "桥架安装", 100])
    workbook.save(workbook_path)

    mappings = [
        {"excel_column_name": "区域", "system_field_name": "area", "field_type": "text"},
        {"excel_column_name": "系统", "system_field_name": "system_name", "field_type": "text"},
        {"excel_column_name": "子项", "system_field_name": "task_name", "field_type": "text"},
        {"excel_column_name": "清单量", "system_field_name": "total_quantity", "field_type": "number"},
        {"excel_column_name": "累计完成", "system_field_name": "cumulative_quantity", "field_type": "number"},
        {"excel_column_name": "本期完成", "system_field_name": "period_quantity", "field_type": "number"},
    ]

    with TestClient(app) as client:
        project_response = client.post("/api/projects", json={"name": "多Sheet真实试用回归"})
        project_id = project_response.json()["id"]
        with workbook_path.open("rb") as file:
            upload_response = client.post(
                f"/api/projects/{project_id}/imports/upload",
                data={"data_date": "2026-05-13"},
                files={"file": (workbook_path.name, file, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            )

        assert upload_response.status_code == 200
        assert upload_response.json()["sheets"] == ["使用说明", "01_标准机电进度表", "02_字段不规范表", "04_异常与合计测试表"]
        batch_id = upload_response.json()["batch"]["id"]

        parse_response = client.post(
            f"/api/imports/{batch_id}/parse",
            json={"sheet_name": "02_字段不规范表", "header_row_index": 1, "data_start_row_index": 2},
        )
        assert parse_response.status_code == 200
        assert [column["name"] for column in parse_response.json()["columns"]] == ["区域", "系统", "子项", "清单量", "累计完成", "本期完成"]
        assert parse_response.json()["preview_rows"][0]["子项"] == "喷淋主管安装"

        missing_sheet_response = client.post(
            f"/api/imports/{batch_id}/parse",
            json={"sheet_name": "不存在的Sheet", "header_row_index": 1, "data_start_row_index": 2},
        )
        assert missing_sheet_response.status_code == 400
        assert missing_sheet_response.json()["detail"]["code"] == "SHEET_NOT_FOUND"

        parse_again_response = client.post(
            f"/api/imports/{batch_id}/parse",
            json={"sheet_name": "02_字段不规范表", "header_row_index": 1, "data_start_row_index": 2},
        )
        assert parse_again_response.status_code == 200

        confirm_response = client.post(
            f"/api/imports/{batch_id}/confirm",
            json={"data_date": "2026-05-13", "import_strategy": "new_batch", "field_mappings": mappings},
        )

    assert confirm_response.status_code == 200
    assert confirm_response.json()["imported_count"] == 1

    db = SessionLocal()
    try:
        batch = db.get(ImportBatch, batch_id)
        assert batch.sheet_name == "02_字段不规范表"
        names = [item.task_name for item in db.query(ProgressItem).filter(ProgressItem.batch_id == batch_id).all()]
        assert names == ["喷淋主管安装"]
        assert "第一个Sheet任务" not in names
    finally:
        db.close()


def test_parse_validate_confirm_multiple_sheets_creates_independent_batches(tmp_path: Path) -> None:
    workbook_path = tmp_path / "multi-import.xlsx"
    workbook = Workbook()
    sheet_a = workbook.active
    sheet_a.title = "机电单位"
    sheet_a.append(["工作内容", "专业", "单位", "总工程量", "累计完成量", "实际完成率"])
    sheet_a.append(["桥架安装", "机电", "米", 100, 60, "60%"])
    sheet_b = workbook.create_sheet("消防单位")
    sheet_b.append(["工作内容", "专业", "单位", "总工程量", "累计完成量", "实际完成率"])
    sheet_b.append(["喷淋安装", "消防", "米", 80, 40, "50%"])
    workbook.save(workbook_path)
    mappings = [
        {"excel_column_name": "工作内容", "system_field_name": "task_name", "field_type": "text"},
        {"excel_column_name": "专业", "system_field_name": "discipline", "field_type": "text"},
        {"excel_column_name": "单位", "system_field_name": "unit", "field_type": "text"},
        {"excel_column_name": "总工程量", "system_field_name": "total_quantity", "field_type": "number"},
        {"excel_column_name": "累计完成量", "system_field_name": "cumulative_quantity", "field_type": "number"},
        {"excel_column_name": "实际完成率", "system_field_name": "actual_percent", "field_type": "percent"},
    ]

    with TestClient(app) as client:
        project_id = client.post("/api/projects", json={"name": "多Sheet独立批次"}).json()["id"]
        db = SessionLocal()
        try:
            baseline = BaselinePlan(project_id=project_id, name="多Sheet基线", is_default=True)
            db.add(baseline)
            db.commit()
            baseline_id = baseline.id
        finally:
            db.close()
        with workbook_path.open("rb") as file:
            upload_response = client.post(
                f"/api/projects/{project_id}/imports/upload",
                data={"data_date": "2026-05-12"},
                files={"file": (workbook_path.name, file, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            )
        file_id = upload_response.json()["batch"]["id"]
        parse_response = client.post(
            f"/api/imports/{file_id}/parse-multiple-sheets",
            json={
                "project_id": project_id,
                "sheet_names": ["机电单位", "消防单位"],
                "header_row_index": 1,
                "data_start_row_index": 2,
                "data_date": "2026-05-12",
                "baseline_plan_id": baseline_id,
            },
        )
        assert parse_response.status_code == 200
        parsed = parse_response.json()
        assert parsed["success_count"] == 2
        batch_ids = [result["batch_id"] for result in parsed["results"]]
        validate_response = client.post(
            "/api/imports/validate-multiple-sheets",
            json={
                "sheets": [
                    {"batch_id": batch_ids[0], "sheet_name": "机电单位", "mappings": mappings},
                    {"batch_id": batch_ids[1], "sheet_name": "消防单位", "mappings": mappings},
                ]
            },
        )
        confirm_response = client.post(
            "/api/imports/confirm-multiple-sheets",
            json={
                "project_id": project_id,
                "data_date": "2026-05-12",
                "baseline_plan_id": baseline_id,
                "sheets": [
                    {"batch_id": batch_ids[0], "sheet_name": "机电单位", "mappings": mappings, "import_strategy": "new_batch"},
                    {"batch_id": batch_ids[1], "sheet_name": "消防单位", "mappings": mappings, "import_strategy": "new_batch"},
                ],
            },
        )
        publish_response = client.post("/api/imports/publish-multiple-sheets", json=batch_ids)

    assert validate_response.status_code == 200
    assert validate_response.json()["success_count"] == 2
    assert confirm_response.status_code == 200
    assert confirm_response.json()["success_count"] == 2
    assert [batch["imported_count"] for batch in confirm_response.json()["batches"]] == [1, 1]
    assert publish_response.status_code == 200
    assert publish_response.json()["published_count"] == 2
    assert publish_response.json()["failed_publish_count"] == 0

    db = SessionLocal()
    try:
        batches = [db.get(ImportBatch, batch_id) for batch_id in batch_ids]
        assert [batch.sheet_name for batch in batches] == ["机电单位", "消防单位"]
        assert batches[0].import_group_id == batches[1].import_group_id
        assert batches[0].import_group_id is not None
        assert batches[0].import_group_name
        assert [batch.data_date.isoformat() for batch in batches] == ["2026-05-12", "2026-05-12"]
        assert [batch.baseline_plan_id for batch in batches] == [baseline_id, baseline_id]
        assert [batch.status for batch in batches] == ["published", "published"]
        assert db.query(ProgressItem).filter(ProgressItem.batch_id.in_(batch_ids)).count() == 2
    finally:
        db.close()

    with TestClient(app) as client:
        list_response = client.get(f"/api/projects/{project_id}/imports")
    assert list_response.status_code == 200
    list_rows = [row for row in list_response.json() if row["id"] in batch_ids]
    assert {row["sheet_name"] for row in list_rows} == {"机电单位", "消防单位"}
    assert {row["baseline_plan_name"] for row in list_rows} == {"多Sheet基线"}
    assert {row["is_multi_sheet"] for row in list_rows} == {True}
    assert {row["group_sheet_count"] for row in list_rows} == {2}


def test_multi_sheet_failure_does_not_block_success_and_replace_same_date_is_sheet_scoped(tmp_path: Path) -> None:
    workbook_path = tmp_path / "multi-import-partial.xlsx"
    workbook = Workbook()
    sheet_a = workbook.active
    sheet_a.title = "机电单位"
    sheet_a.append(["工作内容", "专业", "单位", "总工程量", "累计完成量", "实际完成率"])
    sheet_a.append(["桥架安装", "机电", "米", 100, 60, "60%"])
    sheet_b = workbook.create_sheet("消防单位")
    sheet_b.append(["工作内容", "专业", "单位", "总工程量", "累计完成量", "实际完成率"])
    sheet_b.append(["喷淋安装", "消防", "米", 80, 40, "50%"])
    workbook.save(workbook_path)
    good_mappings = [
        {"excel_column_name": "工作内容", "system_field_name": "task_name", "field_type": "text"},
        {"excel_column_name": "专业", "system_field_name": "discipline", "field_type": "text"},
        {"excel_column_name": "单位", "system_field_name": "unit", "field_type": "text"},
        {"excel_column_name": "总工程量", "system_field_name": "total_quantity", "field_type": "number"},
        {"excel_column_name": "累计完成量", "system_field_name": "cumulative_quantity", "field_type": "number"},
        {"excel_column_name": "实际完成率", "system_field_name": "actual_percent", "field_type": "percent"},
    ]
    blocking_mappings = [
        {"excel_column_name": "工作内容", "system_field_name": "task_name", "field_type": "text"},
        {"excel_column_name": "专业", "system_field_name": "task_name", "field_type": "text"},
    ]

    with TestClient(app) as client:
        project_id = client.post("/api/projects", json={"name": "多Sheet失败隔离"}).json()["id"]
        with workbook_path.open("rb") as file:
            upload_response = client.post(
                f"/api/projects/{project_id}/imports/upload",
                data={"data_date": "2026-05-12"},
                files={"file": (workbook_path.name, file, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            )
        file_id = upload_response.json()["batch"]["id"]
        parse_response = client.post(
            f"/api/imports/{file_id}/parse-multiple-sheets",
            json={
                "project_id": project_id,
                "sheet_names": ["机电单位", "消防单位"],
                "header_row_index": 1,
                "data_start_row_index": 2,
                "data_date": "2026-05-12",
            },
        )
        batch_ids = [result["batch_id"] for result in parse_response.json()["results"]]
        confirm_response = client.post(
            "/api/imports/confirm-multiple-sheets",
            json={
                "project_id": project_id,
                "data_date": "2026-05-12",
                "sheets": [
                    {"batch_id": batch_ids[0], "sheet_name": "机电单位", "mappings": good_mappings, "import_strategy": "replace_same_date"},
                    {"batch_id": batch_ids[1], "sheet_name": "消防单位", "mappings": blocking_mappings, "import_strategy": "replace_same_date"},
                ],
            },
        )
        with workbook_path.open("rb") as file:
            second_upload = client.post(
                f"/api/projects/{project_id}/imports/upload",
                data={"data_date": "2026-05-12"},
                files={"file": (workbook_path.name, file, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            )
        second_file_id = second_upload.json()["batch"]["id"]
        second_parse = client.post(
            f"/api/imports/{second_file_id}/parse-multiple-sheets",
            json={
                "project_id": project_id,
                "sheet_names": ["消防单位"],
                "header_row_index": 1,
                "data_start_row_index": 2,
                "data_date": "2026-05-12",
            },
        )
        replacement_id = second_parse.json()["results"][0]["batch_id"]
        replacement_confirm = client.post(
            "/api/imports/confirm-multiple-sheets",
            json={
                "project_id": project_id,
                "data_date": "2026-05-12",
                "sheets": [
                    {"batch_id": replacement_id, "sheet_name": "消防单位", "mappings": good_mappings, "import_strategy": "replace_same_date"},
                ],
            },
        )

    assert confirm_response.status_code == 200
    assert confirm_response.json()["success_count"] == 1
    assert confirm_response.json()["failed_count"] == 1
    assert replacement_confirm.status_code == 200
    assert replacement_confirm.json()["success_count"] == 1

    db = SessionLocal()
    try:
        first_sheet_batch = db.get(ImportBatch, batch_ids[0])
        failed_sheet_batch = db.get(ImportBatch, batch_ids[1])
        replacement_batch = db.get(ImportBatch, replacement_id)
        assert first_sheet_batch.is_active is True
        assert failed_sheet_batch.is_active is False
        assert replacement_batch.is_active is True
        assert first_sheet_batch.sheet_name == "机电单位"
        assert replacement_batch.sheet_name == "消防单位"
    finally:
        db.close()


def test_multi_sheet_publish_failure_does_not_block_other_batches(tmp_path: Path) -> None:
    workbook_path = tmp_path / "multi-publish-partial.xlsx"
    workbook = Workbook()
    sheet_a = workbook.active
    sheet_a.title = "机电单位"
    sheet_a.append(["工作内容", "专业", "单位", "总工程量", "累计完成量", "实际完成率"])
    sheet_a.append(["桥架安装", "机电", "米", 100, 60, "60%"])
    sheet_b = workbook.create_sheet("消防单位")
    sheet_b.append(["工作内容", "专业", "单位", "总工程量", "累计完成量", "实际完成率"])
    sheet_b.append(["喷淋安装", "消防", "米", 80, 40, "50%"])
    workbook.save(workbook_path)
    mappings = [
        {"excel_column_name": "工作内容", "system_field_name": "task_name", "field_type": "text"},
        {"excel_column_name": "专业", "system_field_name": "discipline", "field_type": "text"},
        {"excel_column_name": "单位", "system_field_name": "unit", "field_type": "text"},
        {"excel_column_name": "总工程量", "system_field_name": "total_quantity", "field_type": "number"},
        {"excel_column_name": "累计完成量", "system_field_name": "cumulative_quantity", "field_type": "number"},
        {"excel_column_name": "实际完成率", "system_field_name": "actual_percent", "field_type": "percent"},
    ]

    with TestClient(app) as client:
        project_id = client.post("/api/projects", json={"name": "批量发布失败隔离"}).json()["id"]
        with workbook_path.open("rb") as file:
            upload_response = client.post(
                f"/api/projects/{project_id}/imports/upload",
                data={"data_date": "2026-05-12"},
                files={"file": (workbook_path.name, file, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            )
        file_id = upload_response.json()["batch"]["id"]
        parse_response = client.post(
            f"/api/imports/{file_id}/parse-multiple-sheets",
            json={
                "project_id": project_id,
                "sheet_names": ["机电单位", "消防单位"],
                "header_row_index": 1,
                "data_start_row_index": 2,
                "data_date": "2026-05-12",
            },
        )
        batch_ids = [result["batch_id"] for result in parse_response.json()["results"]]
        confirm_response = client.post(
            "/api/imports/confirm-multiple-sheets",
            json={
                "project_id": project_id,
                "data_date": "2026-05-12",
                "sheets": [
                    {"batch_id": batch_ids[0], "sheet_name": "机电单位", "mappings": mappings, "import_strategy": "new_batch"},
                ],
            },
        )
        publish_response = client.post("/api/imports/publish-multiple-sheets", json=batch_ids)

    assert confirm_response.status_code == 200
    assert publish_response.status_code == 200
    payload = publish_response.json()
    assert payload["published_count"] == 1
    assert payload["failed_publish_count"] == 1
    result_by_id = {result["batch_id"]: result for result in payload["results"]}
    assert result_by_id[batch_ids[0]]["published"] is True
    assert result_by_id[batch_ids[1]]["published"] is False
    assert result_by_id[batch_ids[1]]["error"] == "只有导入成功的批次可以发布。"

    db = SessionLocal()
    try:
        assert db.get(ImportBatch, batch_ids[0]).status == "published"
        assert db.get(ImportBatch, batch_ids[1]).status == "parsed"
    finally:
        db.close()


def test_multi_sheet_saved_template_records_sheet_name_and_field_structure(tmp_path: Path) -> None:
    workbook_path = tmp_path / "multi-template.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "智能化单位"
    sheet.append(["工作内容", "专业", "单位", "总工程量", "累计完成量", "实际完成率"])
    sheet.append(["弱电桥架安装", "智能化", "米", 50, 20, "40%"])
    workbook.save(workbook_path)
    mappings = [
        {"excel_column_name": "工作内容", "system_field_name": "task_name", "field_type": "text"},
        {"excel_column_name": "专业", "system_field_name": "discipline", "field_type": "text"},
        {"excel_column_name": "单位", "system_field_name": "unit", "field_type": "text"},
        {"excel_column_name": "总工程量", "system_field_name": "total_quantity", "field_type": "number"},
        {"excel_column_name": "累计完成量", "system_field_name": "cumulative_quantity", "field_type": "number"},
        {"excel_column_name": "实际完成率", "system_field_name": "actual_percent", "field_type": "percent"},
    ]

    with TestClient(app) as client:
        project_id = client.post("/api/projects", json={"name": "模板记录Sheet"}).json()["id"]
        with workbook_path.open("rb") as file:
            upload_response = client.post(
                f"/api/projects/{project_id}/imports/upload",
                files={"file": (workbook_path.name, file, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            )
        file_id = upload_response.json()["batch"]["id"]
        parse_response = client.post(
            f"/api/imports/{file_id}/parse-multiple-sheets",
            json={
                "project_id": project_id,
                "sheet_names": ["智能化单位"],
                "header_row_index": 1,
                "data_start_row_index": 2,
            },
        )
        batch_id = parse_response.json()["results"][0]["batch_id"]
        confirm_response = client.post(
            "/api/imports/confirm-multiple-sheets",
            json={
                "project_id": project_id,
                "sheets": [
                    {
                        "batch_id": batch_id,
                        "sheet_name": "智能化单位",
                        "mappings": mappings,
                        "import_strategy": "new_batch",
                        "save_template": True,
                        "template_name": "智能化Sheet模板",
                    }
                ],
            },
        )

    assert confirm_response.status_code == 200
    db = SessionLocal()
    try:
        template = db.query(MappingTemplate).filter(MappingTemplate.name == "智能化Sheet模板").one()
        assert template.sheet_name == "智能化单位"
        field_structure = json.loads(template.field_structure)
        assert field_structure["sheet_name"] == "智能化单位"
        assert [column["excel_column_name"] for column in field_structure["columns"]][:2] == ["工作内容", "专业"]
    finally:
        db.close()


def _upload_and_parse_sample(client: TestClient, project_id: int, data_date: str) -> int:
    sample_path = SAMPLE_DIR / "sample_progress_a.csv"
    with sample_path.open("rb") as file:
        upload_response = client.post(
            f"/api/projects/{project_id}/imports/upload",
            data={"data_date": data_date},
            files={"file": ("sample_progress_a.csv", file, "text/csv")},
        )
    batch_id = upload_response.json()["batch"]["id"]
    parse_response = client.post(
        f"/api/imports/{batch_id}/parse",
        json={"sheet_name": "CSV", "data_date": data_date, "header_row_index": 1, "data_start_row_index": 2},
    )
    assert parse_response.status_code == 200
    return batch_id


def test_parse_accepts_null_header_rows_and_auto_detects(tmp_path: Path) -> None:
    sample_path = tmp_path / "auto-header.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "03_施工单位原始报表"
    sheet.append(["施工单位原始周报（字段名不规范，测试字段别名识别与 extra_fields）"])
    sheet.append(["区域", "系统", "子项", "清单量", "累计完成", "本期完成", "应完成率", "进度百分比"])
    sheet.append(["地下室", "消防系统", "喷淋主管安装", 300, 120, 30, "50%", "40%"])
    workbook.save(sample_path)

    with TestClient(app) as client:
        project_id = client.post("/api/projects", json={"name": "T007 自动识别"}).json()["id"]
        with sample_path.open("rb") as file:
            upload_response = client.post(
                f"/api/projects/{project_id}/imports/upload",
                files={"file": ("auto-header.xlsx", file, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            )
        batch_id = upload_response.json()["batch"]["id"]

        parse_response = client.post(
            f"/api/imports/{batch_id}/parse",
            json={
                "sheet_name": "03_施工单位原始报表",
                "header_row_index": None,
                "data_start_row_index": None,
                "multi_header": False,
                "header_end_row_index": None,
            },
        )

        assert parse_response.status_code == 200
        parsed = parse_response.json()
        assert parsed["batch"]["header_row_index"] == 2
        assert parsed["batch"]["data_start_row_index"] == 3
        assert [column["name"] for column in parsed["columns"][:3]] == ["区域", "系统", "子项"]


def test_parse_keeps_manual_header_rows_when_integers_are_sent(tmp_path: Path) -> None:
    sample_path = tmp_path / "manual-header.csv"
    sample_path.write_text("说明行,,\n区域,系统,子项\n地下室,消防系统,喷淋主管安装\n", encoding="utf-8-sig")

    with TestClient(app) as client:
        project_id = client.post("/api/projects", json={"name": "T007 手动设置"}).json()["id"]
        with sample_path.open("rb") as file:
            upload_response = client.post(
                f"/api/projects/{project_id}/imports/upload",
                files={"file": ("manual-header.csv", file, "text/csv")},
            )
        batch_id = upload_response.json()["batch"]["id"]

        parse_response = client.post(
            f"/api/imports/{batch_id}/parse",
            json={
                "sheet_name": "CSV",
                "header_row_index": 2,
                "data_start_row_index": 3,
                "multi_header": False,
                "header_end_row_index": None,
            },
        )

        assert parse_response.status_code == 200
        parsed = parse_response.json()
        assert parsed["batch"]["header_row_index"] == 2
        assert parsed["batch"]["data_start_row_index"] == 3
        assert parsed["preview_rows"][0]["子项"] == "喷淋主管安装"


def _import_and_publish_sample(client: TestClient, project_id: int, data_date: str, strategy: str) -> int:
    batch_id = _upload_and_parse_sample(client, project_id, data_date)
    confirm_response = client.post(
        f"/api/imports/{batch_id}/confirm",
        json={
            "data_date": data_date,
            "import_strategy": strategy,
            "field_mappings": FIELD_MAPPINGS_A,
        },
    )
    assert confirm_response.status_code == 200
    publish_response = client.post(f"/api/imports/{batch_id}/publish")
    assert publish_response.status_code == 200
    return batch_id
