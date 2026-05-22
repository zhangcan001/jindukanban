from pathlib import Path

from openpyxl import Workbook

import pytest

from app.services.excel_parser import ExcelParseError, detect_header_row, get_sheet_names, parse_preview, parse_rows, recommend_header_rows


def test_csv_can_parse_and_detect_header(tmp_path: Path) -> None:
    path = tmp_path / "progress.csv"
    path.write_text("说明行,,\nWBS编码,工作内容,累计完成量\nJD.01,桥架安装,50\n", encoding="utf-8-sig")

    columns, rows, row_count = parse_preview(str(path), "CSV", 2, 3, False, None)

    assert get_sheet_names(str(path)) == ["CSV"]
    assert row_count == 1
    assert columns[0]["recommended_field"] == "wbs_code"
    assert rows[0]["工作内容"] == "桥架安装"


def test_xlsx_can_parse_with_multi_header(tmp_path: Path) -> None:
    path = tmp_path / "progress.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "进度"
    sheet.append(["任务", "任务", "进度"])
    sheet.append(["WBS编码", "工作内容", "实际完成率"])
    sheet.append(["JD.01", "桥架安装", "58%"])
    workbook.save(path)

    columns, rows, row_count = parse_preview(str(path), "进度", 1, 3, True, 2)
    parsed_rows = parse_rows(str(path), "进度", 1, 3, True, 2)

    assert get_sheet_names(str(path)) == ["进度"]
    assert row_count == 1
    assert columns[0]["name"] == "任务_WBS编码"
    assert rows[0]["任务_工作内容"] == "桥架安装"
    assert parsed_rows[0]["进度_实际完成率"] == "58%"


def test_xlsx_merged_header_cells_are_filled_and_combined(tmp_path: Path) -> None:
    path = tmp_path / "merged-header.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "进度"
    sheet.merge_cells("A1:A2")
    sheet.merge_cells("B1:C1")
    sheet.merge_cells("D1:E1")
    sheet["A1"] = "工作内容"
    sheet["B1"] = "计划"
    sheet["D1"] = "实际"
    sheet["B2"] = "完成率"
    sheet["C2"] = "完成量"
    sheet["D2"] = "完成率"
    sheet["E2"] = "完成量"
    sheet.append(["桥架安装", "60%", 60, "50%", 50])
    workbook.save(path)

    columns, rows, row_count = parse_preview(str(path), "进度", 1, 3, True, 2)
    parsed_rows = parse_rows(str(path), "进度", 1, 3, True, 2)

    assert row_count == 1
    assert [column["name"] for column in columns] == ["工作内容", "计划完成率", "计划完成量", "实际完成率", "实际完成量"]
    assert rows[0]["计划完成率"] == "60%"
    assert parsed_rows[0]["实际完成率"] == "50%"


def test_xlsx_merged_header_can_be_inferred_without_multi_header_flag(tmp_path: Path) -> None:
    path = tmp_path / "merged-header-auto.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "06_合并表头样例"
    sheet.merge_cells("G1:J1")
    sheet.merge_cells("K1:L1")
    sheet.append(["楼栋", "楼层", "专业", "系统", "工作内容", "单位", "工程量", "", "", "", "完成率", ""])
    sheet.append(["", "", "", "", "", "", "总工程量", "计划完成量", "累计完成量", "本期完成量", "计划完成率", "实际完成率"])
    sheet.append(["1号楼", "1层", "机电", "给排水", "给水管安装", "米", 100, 60, 50, 10, "60%", "50%"])
    workbook.save(path)

    columns, rows, row_count = parse_preview(str(path), "06_合并表头样例", 1, None, False, None)

    detected = {column["recommended_field"] for column in columns}
    assert row_count == 1
    assert rows[0]["楼栋"] == "1号楼"
    assert {
        "building",
        "floor",
        "discipline",
        "system_name",
        "task_name",
        "unit",
        "total_quantity",
        "planned_quantity",
        "cumulative_quantity",
        "period_quantity",
        "planned_percent",
        "actual_percent",
    }.issubset(detected)


def test_parent_header_row_before_detected_header_is_combined(tmp_path: Path) -> None:
    path = tmp_path / "merged-parent-before-header.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "06_合并表头样例"
    sheet.merge_cells("A1:O1")
    sheet["A1"] = "06 合并表头 / 多行表头样例"
    sheet.merge_cells("A2:C2")
    sheet.merge_cells("D2:F2")
    sheet.merge_cells("G2:I2")
    sheet.merge_cells("J2:K2")
    sheet.merge_cells("L2:M2")
    sheet.merge_cells("N2:O2")
    sheet.append(["工程信息", "", "", "任务信息", "", "", "工程量", "", "", "计划", "", "实际", "", "时间", ""])
    sheet.append(["楼栋", "楼层", "专业", "系统", "工作内容", "单位", "总工程量", "计划完成量", "累计完成量", "完成率", "完成量", "完成率", "本期完成量", "计划开始", "计划完成"])
    sheet.append(["1号楼", "B1层", "电气", "桥架系统", "主干桥架安装", "米", 520, 482, 440, 0.927, 482, 0.846, 35, "2026-05-01", "2026-06-10"])
    workbook.save(path)

    columns, rows, row_count = parse_preview(str(path), "06_合并表头样例", None, None, False, None)

    detected = {column["name"]: column["recommended_field"] for column in columns}
    assert row_count == 1
    assert rows[0]["工程信息_楼栋"] == "1号楼"
    assert detected["任务信息_工作内容"] == "task_name"
    assert detected["计划完成率"] == "planned_percent"
    assert detected["计划完成量"] == "planned_quantity"
    assert detected["实际完成率"] == "actual_percent"
    assert detected["实际本期完成量"] == "period_quantity"


def test_multi_sheet_xlsx_returns_all_sheet_names_and_parses_selected_sheet(tmp_path: Path) -> None:
    path = tmp_path / "multi-sheet.xlsx"
    workbook = Workbook()
    workbook.active.title = "使用说明"
    workbook.active.append(["说明"])
    standard = workbook.create_sheet("01_标准机电进度表")
    standard.append(["WBS编码", "工作内容", "实际完成率"])
    standard.append(["JD.01", "桥架安装", "70%"])
    irregular = workbook.create_sheet("02_字段不规范表")
    irregular.append(["区域", "系统", "子项", "清单量", "累计完成", "本期完成"])
    irregular.append(["地下室", "消防系统", "喷淋主管安装", 300, 120, 30])
    workbook.save(path)

    columns, rows, row_count = parse_preview(str(path), "02_字段不规范表", 1, 2, False, None)

    assert get_sheet_names(str(path)) == ["使用说明", "01_标准机电进度表", "02_字段不规范表"]
    assert row_count == 1
    assert [column["name"] for column in columns] == ["区域", "系统", "子项", "清单量", "累计完成", "本期完成"]
    assert rows[0]["子项"] == "喷淋主管安装"


def test_parse_missing_sheet_raises_sheet_not_found(tmp_path: Path) -> None:
    path = tmp_path / "multi-sheet.xlsx"
    workbook = Workbook()
    workbook.active.title = "存在的Sheet"
    workbook.active.append(["工作内容"])
    workbook.active.append(["桥架安装"])
    workbook.save(path)

    with pytest.raises(ExcelParseError) as exc_info:
        parse_preview(str(path), "不存在的Sheet", 1, 2, False, None)

    assert exc_info.value.code == "SHEET_NOT_FOUND"


def test_auto_detects_second_row_when_first_row_is_title(tmp_path: Path) -> None:
    path = tmp_path / "title-row.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "03_施工单位原始报表"
    sheet.append(["施工单位原始周报（字段名不规范，测试字段别名识别与 extra_fields）"])
    sheet.append(["区域", "系统", "子项", "清单量", "累计完成", "本期完成", "应完成率", "进度百分比", "责任人", "班组", "材料状态", "施工备注"])
    sheet.append(["地下室", "消防系统", "喷淋主管安装", 300, 120, 30, "50%", "40%", "张三", "一班", "已到场", "正常"])
    workbook.save(path)

    columns, rows, row_count = parse_preview(str(path), "03_施工单位原始报表", None, None, False, None)

    assert row_count == 1
    assert [column["name"] for column in columns] == [
        "区域",
        "系统",
        "子项",
        "清单量",
        "累计完成",
        "本期完成",
        "应完成率",
        "进度百分比",
        "责任人",
        "班组",
        "材料状态",
        "施工备注",
    ]
    assert rows[0]["子项"] == "喷淋主管安装"


def test_auto_detects_first_row_when_first_row_is_header(tmp_path: Path) -> None:
    path = tmp_path / "first-row-header.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "进度"
    sheet.append(["区域", "系统", "子项", "清单量", "累计完成", "本期完成"])
    sheet.append(["地下室", "消防系统", "喷淋主管安装", 300, 120, 30])
    workbook.save(path)

    columns, rows, row_count = parse_preview(str(path), "进度", None, None, False, None)

    assert row_count == 1
    assert columns[0]["name"] == "区域"
    assert rows[0]["区域"] == "地下室"


def test_long_contractor_weekly_report_title_is_not_header(tmp_path: Path) -> None:
    path = tmp_path / "contractor-title.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "03_施工单位原始报表"
    sheet.append(["施工单位原始周报（字段名不规范，测试字段别名识别与 extra_fields）"])
    sheet.append(["区域", "系统", "子项", "清单量"])
    sheet.append(["地下室", "消防系统", "喷淋主管安装", 300])
    workbook.save(path)

    columns, _, _ = parse_preview(str(path), "03_施工单位原始报表", None, None, False, None)

    assert columns[0]["name"] != "施工单位原始周报（字段名不规范，测试字段别名识别与 extra_fields）"
    assert columns[0]["name"] == "区域"


def test_t006_contractor_columns_detect_expected_system_fields(tmp_path: Path) -> None:
    path = tmp_path / "contractor-columns.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "03_施工单位原始报表"
    sheet.append(["施工单位原始周报（字段名不规范，测试字段别名识别与 extra_fields）"])
    sheet.append(["区域", "系统", "子项", "清单量", "累计完成", "本期完成", "应完成率", "进度百分比", "责任人", "班组", "材料状态", "施工备注"])
    sheet.append(["地下室", "消防系统", "喷淋主管安装", 300, 120, 30, "50%", "40%", "张三", "一班", "已到场", "正常"])
    workbook.save(path)

    columns, _, _ = parse_preview(str(path), "03_施工单位原始报表", None, None, False, None)
    detected = {column["name"]: column["recommended_field"] for column in columns}

    assert detected["区域"] == "area"
    assert detected["系统"] == "system_name"
    assert detected["子项"] == "task_name"
    assert detected["清单量"] == "total_quantity"
    assert detected["累计完成"] == "cumulative_quantity"
    assert detected["本期完成"] == "period_quantity"
    assert detected["应完成率"] == "planned_percent"
    assert detected["进度百分比"] == "actual_percent"


def test_detect_header_row_returns_one_based_second_row_for_title_sheet() -> None:
    import pandas as pd

    raw = pd.DataFrame(
        [
            ["施工单位原始周报（字段名不规范，测试字段别名识别与 extra_fields）", "", "", ""],
            ["区域", "系统", "子项", "清单量"],
            ["地下室", "消防系统", "喷淋主管安装", 300],
        ]
    )

    assert detect_header_row(raw) == 2


def test_recommend_header_rows_scans_first_twenty_rows() -> None:
    import pandas as pd

    raw = pd.DataFrame(
        [["说明", "", "", ""] for _ in range(17)]
        + [
            ["区域", "系统", "子项", "清单量"],
            ["地下室", "消防系统", "喷淋主管安装", 300],
        ]
    )

    recommendation = recommend_header_rows(raw)

    assert recommendation["header_row_index"] == 18
    assert recommendation["data_start_row_index"] == 19
    assert recommendation["confidence"] in {"高", "中"}
