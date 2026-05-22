from datetime import date, timedelta
from io import BytesIO

from docx import Document
from fastapi.testclient import TestClient
from openpyxl import load_workbook

from app.database import SessionLocal
from app.main import app
from app.models.baseline_plan import BaselinePlan
from app.models.import_batch import ImportBatch
from app.models.progress_item import ProgressItem
from app.models.project import Project
from app.models.rectification_item import RectificationItem
from app.models.warning_record import WarningRecord


def _seed_rectifications():
    db = SessionLocal()
    try:
        project = Project(name="整改 beta 项目", project_type="测试")
        db.add(project)
        db.flush()
        baseline = BaselinePlan(project_id=project.id, name="整改来源计划", is_default=True)
        db.add(baseline)
        db.flush()
        batch = ImportBatch(project_id=project.id, file_name="beta.csv", status="published", data_date=date(2026, 5, 14), imported_count=3, baseline_plan_id=baseline.id)
        db.add(batch)
        db.flush()
        item_a = ProgressItem(
            project_id=project.id,
            batch_id=batch.id,
            task_id=1,
            task_name="喷淋主管安装",
            discipline="消防",
            building="A1",
            floor="1层",
            system_name="喷淋",
            actual_percent=40,
            planned_percent=70,
            progress_deviation=-30,
            status="seriously_delayed",
            planned_start_date=date(2026, 5, 1),
            planned_finish_date=date(2026, 5, 15),
        )
        item_b = ProgressItem(
            project_id=project.id,
            batch_id=batch.id,
            task_id=2,
            task_name="消火栓箱安装",
            discipline="消防",
            building="A1",
            floor="2层",
            system_name="消火栓",
            actual_percent=60,
            planned_percent=70,
            progress_deviation=-10,
            status="delayed",
            planned_start_date=date(2026, 5, 1),
            planned_finish_date=date(2026, 5, 16),
        )
        db.add_all([item_a, item_b])
        db.flush()
        warning = WarningRecord(project_id=project.id, batch_id=batch.id, task_id=item_b.task_id, level="warning", title="进度预警", message="消火栓滞后")
        db.add(warning)
        db.flush()
        db.add_all(
            [
                RectificationItem(
                    project_id=project.id,
                    batch_id=batch.id,
                    progress_item_id=item_a.id,
                    source_type="progress_item",
                    source_id=item_a.id,
                    discipline="消防",
                    building="A1",
                    floor="1层",
                    system_name="喷淋",
                    task_name="喷淋主管安装",
                    issue_description="严重滞后",
                    delay_level="seriously_delayed",
                    progress_deviation=-30,
                    responsible_person="张工",
                    planned_finish_date=date.today() - timedelta(days=1),
                    status="open",
                ),
                RectificationItem(
                    project_id=project.id,
                    batch_id=batch.id,
                    warning_record_id=warning.id,
                    source_type="warning",
                    source_id=warning.id,
                    discipline="消防",
                    building="A1",
                    floor="2层",
                    system_name="消火栓",
                    task_name="消火栓箱安装",
                    issue_description="一般滞后",
                    delay_level="delayed",
                    responsible_unit="消防分包",
                    status="in_progress",
                ),
                RectificationItem(project_id=project.id, batch_id=batch.id, source_type="manual", task_name="资料补录", status="closed"),
            ]
        )
        db.commit()
        return project.id, batch.id, item_a.id, warning.id
    finally:
        db.close()


def test_rectification_list_filters_pagination_summary_batch_logs_and_export() -> None:
    project_id, batch_id, progress_item_id, warning_id = _seed_rectifications()

    with TestClient(app) as client:
        response = client.get(f"/api/projects/{project_id}/rectifications", params={"page": 1, "page_size": 2})
        assert response.status_code == 200
        payload = response.json()
        assert payload["total"] == 3
        assert len(payload["items"]) == 2

        assert client.get(f"/api/projects/{project_id}/rectifications", params={"status": "open"}).json()["total"] == 1
        assert client.get(f"/api/projects/{project_id}/rectifications", params={"building": "A1"}).json()["total"] == 2
        assert client.get(f"/api/projects/{project_id}/rectifications", params={"floor": "1层"}).json()["total"] == 1
        assert client.get(f"/api/projects/{project_id}/rectifications", params={"responsible_person": "张工"}).json()["total"] == 1
        assert client.get(f"/api/projects/{project_id}/rectifications", params={"overdue": True}).json()["total"] == 1

        summary = client.get(f"/api/projects/{project_id}/rectifications/summary").json()
        assert summary["total"] == 3
        assert summary["open"] == 1
        assert summary["in_progress"] == 1
        assert summary["closed"] == 1
        assert summary["overdue"] == 1
        assert summary["serious"] == 1

        first_id = payload["items"][0]["id"]
        batch_response = client.post(
            f"/api/projects/{project_id}/rectifications/batch-update",
            json={"ids": [first_id], "status": "completed", "responsible_person": "李工", "planned_finish_date": str(date.today())},
        )
        assert batch_response.status_code == 200
        assert batch_response.json()["updated_count"] == 1

        logs = client.get(f"/api/projects/{project_id}/rectifications/{first_id}/logs").json()
        assert logs
        assert logs[0]["operator"] == "本地用户"

        duplicate_a = client.post(
            f"/api/projects/{project_id}/rectifications/from-progress-items",
            json={"batch_id": batch_id, "progress_item_id": progress_item_id},
        ).json()
        assert duplicate_a["created"] is False

        duplicate_b = client.post(f"/api/projects/{project_id}/rectifications/from-warnings", json={"warning_record_id": warning_id}).json()
        assert duplicate_b["created"] is False

        export_response = client.get(f"/api/projects/{project_id}/rectifications/export", params={"responsible_person": "李工"})
        assert export_response.status_code == 200
        workbook = load_workbook(BytesIO(export_response.content), read_only=True)
        try:
            rows = list(workbook["整改跟踪表"].iter_rows(values_only=True))
            values = [cell for row in rows for cell in row]
            assert "筛选责任人" in values
            assert "李工" in values
            assert "张工" not in values
            assert "批次绑定计划基线" in values
            assert "当前查看计划基线" in values
            assert "整改来源计划" in values
            header = next(row for row in rows if row and row[0] == "状态")
            assert "整改记录摘要" in header
            assert "来源计划基线" in header
            assert "当前统计口径" in values
            assert "统计口径说明" in values
        finally:
            workbook.close()


def test_rectification_report_integrations_include_beta_summary_and_detail() -> None:
    project_id, batch_id, _, _ = _seed_rectifications()

    with TestClient(app) as client:
        dashboard = client.get(f"/api/projects/{project_id}/reports/dashboard-export", params={"batch_id": batch_id})
        assert dashboard.status_code == 200
        workbook = load_workbook(BytesIO(dashboard.content), read_only=True)
        try:
            assert "整改闭环摘要" in workbook.sheetnames
            assert "整改项明细" in workbook.sheetnames
            detail_values = [cell for row in workbook["整改项明细"].iter_rows(values_only=True) for cell in row]
            assert "是否逾期" in detail_values
            assert "喷淋主管安装" in detail_values
        finally:
            workbook.close()

        weekly = client.get(f"/api/projects/{project_id}/reports/weekly-word", params={"batch_id": batch_id})
        assert weekly.status_code == 200
        document = Document(BytesIO(weekly.content))
        text = "\n".join(paragraph.text for paragraph in document.paragraphs)
        assert "整改闭环摘要" in text
        assert "当前项目共有 3 条整改项" in text


def test_empty_rectification_summary_list_and_logs_are_stable() -> None:
    db = SessionLocal()
    try:
        project = Project(name="空整改项目", project_type="测试")
        db.add(project)
        db.commit()
        project_id = project.id
    finally:
        db.close()

    with TestClient(app) as client:
        summary = client.get(f"/api/projects/{project_id}/rectifications/summary")
        listing = client.get(f"/api/projects/{project_id}/rectifications")

    assert summary.status_code == 200
    assert summary.json() == {
        "total": 0,
        "open": 0,
        "in_progress": 0,
        "completed": 0,
        "closed": 0,
        "ignored": 0,
        "overdue": 0,
        "serious": 0,
        "new_this_week": 0,
        "closed_this_week": 0,
    }
    assert listing.status_code == 200
    assert listing.json()["items"] == []
    assert listing.json()["total"] == 0
