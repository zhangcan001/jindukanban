import io
import json
from datetime import date

from openpyxl import load_workbook

from app.database import SessionLocal
from app.models.import_batch import ImportBatch
from app.models.import_validation_issue import ImportValidationIssue
from app.models.project import Project
from app.models.raw_import_row import RawImportRow
from app.services.import_error_report_service import build_error_report_workbook


def _make_batch_with_issues() -> tuple[int, int]:
    db = SessionLocal()
    try:
        project = Project(name="错误清单测试项目")
        db.add(project)
        db.flush()
        batch = ImportBatch(
            project_id=project.id,
            file_name="进度.xlsx",
            sheet_name="电气进度",
            data_date=date(2026, 5, 16),
            row_count=2,
            error_count=2,
            warning_count=1,
            status="parsed",
            is_active=True,
        )
        db.add(batch)
        db.flush()
        db.add_all(
            [
                RawImportRow(
                    batch_id=batch.id,
                    row_index=1,
                    raw_data=json.dumps({"任务名称": "桥架安装", "实际完成率": "abc"}, ensure_ascii=False),
                ),
                RawImportRow(
                    batch_id=batch.id,
                    row_index=2,
                    raw_data=json.dumps({"任务名称": "", "计划完成日期": "2026-04-01"}, ensure_ascii=False),
                ),
            ]
        )
        db.add_all(
            [
                ImportValidationIssue(
                    batch_id=batch.id,
                    row_index=1,
                    column_name="实际完成率",
                    level="error",
                    code="invalid_percent",
                    message="百分比字段无法解析。",
                ),
                ImportValidationIssue(
                    batch_id=batch.id,
                    row_index=2,
                    column_name=None,
                    level="error",
                    code="task_name_empty",
                    message="任务名称不能为空。",
                ),
                ImportValidationIssue(
                    batch_id=batch.id,
                    row_index=2,
                    column_name="计划完成日期",
                    level="warning",
                    code="planned_finish_date_missing",
                    message="缺少计划完成时间。",
                ),
            ]
        )
        db.commit()
        return project.id, batch.id
    finally:
        db.close()


def test_build_error_report_workbook_writes_each_issue_as_row() -> None:
    _, batch_id = _make_batch_with_issues()
    db = SessionLocal()
    try:
        batch = db.get(ImportBatch, batch_id)
        assert batch is not None
        content = build_error_report_workbook(db, batch)
    finally:
        db.close()

    workbook = load_workbook(io.BytesIO(content))
    assert "导入错误清单" in workbook.sheetnames
    sheet = workbook["导入错误清单"]
    rows = list(sheet.iter_rows(values_only=True))
    assert rows[0] == ("行号", "级别", "错误码", "错误说明", "列名", "原始值", "整行原始数据")
    # 3 个 issue → 3 行数据
    data_rows = rows[1:]
    assert len(data_rows) == 3
    # 行号、级别、错误码顺序按 row_index 升序、id 升序
    assert data_rows[0][:5] == (1, "错误", "invalid_percent", "百分比字段无法解析。", "实际完成率")
    assert data_rows[0][5] == "abc"  # 原始值正确回填
    assert data_rows[1][:3] == (2, "错误", "task_name_empty")
    assert data_rows[2][:3] == (2, "警告", "planned_finish_date_missing")

    # 第二个 sheet 包含批次信息
    summary_sheet = workbook["批次信息"]
    summary_values = {row[0]: row[1] for row in summary_sheet.iter_rows(values_only=True)}
    assert summary_values["Sheet"] == "电气进度"
    assert summary_values["错误数"] == 2
    assert summary_values["警告数"] == 1


def test_build_error_report_workbook_handles_empty_issues() -> None:
    db = SessionLocal()
    try:
        project = Project(name="无错误批次")
        db.add(project)
        db.flush()
        batch = ImportBatch(
            project_id=project.id,
            file_name="ok.xlsx",
            sheet_name="给排水",
            status="validated",
            is_active=True,
        )
        db.add(batch)
        db.commit()
        db.refresh(batch)
        content = build_error_report_workbook(db, batch)
    finally:
        db.close()

    workbook = load_workbook(io.BytesIO(content))
    sheet = workbook["导入错误清单"]
    rows = list(sheet.iter_rows(values_only=True))
    assert len(rows) == 2  # header + "无校验问题。" 占位行
    assert "无校验问题" in (rows[1][3] or "")
