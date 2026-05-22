from datetime import date
from uuid import uuid4

from sqlalchemy import delete
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from io import BytesIO

from app.database import SessionLocal, init_db
from app.main import app
from app.models.import_batch import ImportBatch
from app.models.progress_item import ProgressItem
from app.models.progress_task import ProgressTask
from app.models.project import Project
from app.models.warning_record import WarningRecord
from app.models.warning_rule import WarningRule
from app.services.warning_service import run_warning_rules


def test_run_warning_rules_creates_task_records_but_not_quality_records() -> None:
    init_db()
    project_name = f"pytest-warning-{uuid4()}"
    db = SessionLocal()
    try:
        project = Project(name=project_name, project_type="测试")
        db.add(project)
        db.flush()

        task = ProgressTask(project_id=project.id, task_name="桥架安装", normalized_task_name="桥架安装")
        db.add(task)
        db.flush()

        previous_batch = ImportBatch(
            project_id=project.id,
            file_name="previous.csv",
            status="published",
            data_date=date(2026, 5, 6),
        )
        current_batch = ImportBatch(
            project_id=project.id,
            file_name="current.csv",
            status="published",
            data_date=date(2026, 5, 13),
            data_quality_score=60,
        )
        db.add_all([previous_batch, current_batch])
        db.flush()

        previous_item = ProgressItem(
            project_id=project.id,
            batch_id=previous_batch.id,
            task_id=task.id,
            task_name="桥架安装",
                actual_percent=40,
                planned_percent=55,
                progress_deviation=-15,
                current_period_quantity=0,
                planned_start_date=date(2026, 5, 1),
                planned_finish_date=date(2026, 5, 12),
            )
        current_item = ProgressItem(
            project_id=project.id,
            batch_id=current_batch.id,
            task_id=task.id,
            task_name="桥架安装",
            building="A1",
            floor="3层",
            discipline="消防",
            system_name="喷淋系统",
            actual_percent=40,
            planned_percent=60,
                progress_deviation=-20,
                current_period_quantity=0,
                planned_start_date=date(2026, 5, 1),
                planned_finish_date=date(2026, 5, 15),
            )
        db.add_all([previous_item, current_item])
        db.flush()

        records = run_warning_rules(db, project.id, current_batch)

        assert not any("数据质量评分" in (record.title or "") or "数据质量评分" in (record.message or "") for record in records)
        assert any(record.task_id == task.id and record.level in {"warning", "critical"} for record in records)
        assert any("A1" in (record.message or "") and "3层" in (record.message or "") for record in records)
        assert len(records) >= 3
    finally:
        project_id = project.id if "project" in locals() and project.id else None
        if project_id is not None:
            db.execute(delete(WarningRecord).where(WarningRecord.project_id == project_id))
            db.execute(delete(WarningRule).where(WarningRule.project_id == project_id))
            db.execute(delete(ProgressItem).where(ProgressItem.project_id == project_id))
            db.execute(delete(ImportBatch).where(ImportBatch.project_id == project_id))
            db.execute(delete(ProgressTask).where(ProgressTask.project_id == project_id))
            db.execute(delete(Project).where(Project.id == project_id))
            db.commit()
        db.close()


def test_run_warning_rules_skips_serious_delay_before_plan_start() -> None:
    init_db()
    db = SessionLocal()
    try:
        project = Project(name=f"pytest-warning-plan-start-{uuid4()}", project_type="测试")
        db.add(project)
        db.flush()
        rule = WarningRule(project_id=project.id, name="严重滞后超过 10%", rule_type="serious_delay", level="critical", threshold_value=10)
        task = ProgressTask(project_id=project.id, task_name="桥架安装", normalized_task_name="桥架安装")
        batch = ImportBatch(project_id=project.id, file_name="current.csv", status="published", data_date=date(2026, 5, 18))
        db.add_all([rule, task, batch])
        db.flush()
        item = ProgressItem(
            project_id=project.id,
            batch_id=batch.id,
            task_id=task.id,
            task_name="桥架安装",
            actual_percent=0,
            planned_percent=40,
            progress_deviation=-40,
            planned_start_date=date(2026, 5, 20),
            status="not_started_by_plan",
        )
        db.add(item)
        db.flush()

        records = run_warning_rules(db, project.id, batch)

        assert not any(record.title and "严重滞后" in record.title for record in records)
    finally:
        project_id = project.id if "project" in locals() and project.id else None
        if project_id is not None:
            db.execute(delete(WarningRecord).where(WarningRecord.project_id == project_id))
            db.execute(delete(WarningRule).where(WarningRule.project_id == project_id))
            db.execute(delete(ProgressItem).where(ProgressItem.project_id == project_id))
            db.execute(delete(ImportBatch).where(ImportBatch.project_id == project_id))
            db.execute(delete(ProgressTask).where(ProgressTask.project_id == project_id))
            db.execute(delete(Project).where(Project.id == project_id))
            db.commit()
        db.close()


def test_warning_records_include_location_fields_filters_and_export() -> None:
    init_db()
    db = SessionLocal()
    try:
        project = Project(name=f"pytest-warning-location-{uuid4()}")
        db.add(project)
        db.flush()
        rule = WarningRule(project_id=project.id, name="严重滞后超过 10%", rule_type="serious_delay", level="critical", threshold_value=10)
        task = ProgressTask(project_id=project.id, task_name="喷淋主管安装", normalized_task_name="喷淋主管安装")
        db.add_all([rule, task])
        db.flush()
        batch = ImportBatch(project_id=project.id, file_name="current.csv", status="published", data_date=date(2026, 5, 12))
        db.add(batch)
        db.flush()
        item = ProgressItem(
            project_id=project.id,
            batch_id=batch.id,
            task_id=task.id,
            task_name="喷淋主管安装",
            building="A1",
            floor="3层",
            discipline="消防",
            system_name="喷淋系统",
            unit="%",
            actual_percent=58,
            planned_percent=69,
            progress_deviation=-11,
        )
        missing_location_item = ProgressItem(
            project_id=project.id,
            batch_id=batch.id,
            task_name="",
            actual_percent=20,
            planned_percent=40,
            progress_deviation=-20,
        )
        db.add_all([item, missing_location_item])
        db.flush()
        db.add_all(
            [
                WarningRecord(
                    project_id=project.id,
                    batch_id=batch.id,
                    task_id=task.id,
                    rule_id=rule.id,
                    level="critical",
                    title="喷淋主管安装 严重滞后",
                    message="严重滞后",
                ),
                WarningRecord(
                    project_id=project.id,
                    batch_id=batch.id,
                    task_id=None,
                    rule_id=rule.id,
                    level="warning",
                    title="旧预警",
                    message="旧预警没有进度项",
                ),
            ]
        )
        db.commit()
        project_id = project.id
        batch_id = batch.id
    finally:
        db.close()

    with TestClient(app) as client:
        response = client.get(f"/api/projects/{project_id}/warnings", params={"batch_id": batch_id})
        exported = client.get(f"/api/projects/{project_id}/warnings/export", params={"batch_id": batch_id})

    assert response.status_code == 200
    payload = response.json()
    assert len(payload) == 2
    assert any(row["task_name"] == "喷淋主管安装" and row["building"] == "A1" and row["floor"] == "3层" for row in payload)
    assert any(row["task_name"] == "未填写施工项" and row["building"] == "未填写楼栋" for row in payload)
    assert exported.status_code == 200
    workbook = load_workbook(BytesIO(exported.content), read_only=True)
    try:
        text = "\n".join(str(cell) for row in workbook.active.iter_rows(values_only=True) for cell in row if cell is not None)
        assert "喷淋主管安装" in text
        assert "旧预警没有进度项" in text
    finally:
        workbook.close()


def test_legacy_data_quality_warning_records_are_filtered_and_not_rectifiable() -> None:
    init_db()
    db = SessionLocal()
    try:
        project = Project(name=f"pytest-warning-quality-legacy-{uuid4()}")
        db.add(project)
        db.flush()
        batch = ImportBatch(project_id=project.id, file_name="quality.csv", status="published", data_date=date(2026, 5, 18), data_quality_score=65)
        rule = WarningRule(project_id=project.id, name="数据质量评分低于 70", rule_type="low_data_quality", level="warning", threshold_value=70)
        db.add_all([batch, rule])
        db.flush()
        db.add(
            WarningRecord(
                project_id=project.id,
                batch_id=batch.id,
                task_id=None,
                rule_id=rule.id,
                level="warning",
                title="数据质量评分偏低",
                message="当前数据质量评分 65.0，低于阈值 70.0，请检查导入数据质量。",
            )
        )
        db.commit()
        project_id = project.id
        batch_id = batch.id
        warning_id = db.query(WarningRecord.id).filter(WarningRecord.project_id == project_id).scalar()
    finally:
        db.close()

    with TestClient(app) as client:
        records = client.get(f"/api/projects/{project_id}/warnings", params={"batch_id": batch_id})
        create_rectification = client.post(f"/api/projects/{project_id}/rectifications/from-warnings", json={"warning_record_id": warning_id})
        run = client.post(f"/api/projects/{project_id}/warnings/run", params={"batch_id": batch_id})

    assert records.status_code == 200
    assert records.json() == []
    assert run.status_code == 200
    assert run.json()["generated_count"] == 0
    assert run.json()["records"] == []
    assert create_rectification.status_code == 400
    assert "数据质量评分低" in create_rectification.text
    db = SessionLocal()
    try:
        legacy = db.get(WarningRecord, warning_id)
        assert legacy is not None
        assert "数据质量评分" in (legacy.message or "")
    finally:
        db.close()


def test_quality_rule_is_not_created_for_new_warning_rules() -> None:
    init_db()
    db = SessionLocal()
    try:
        project = Project(name=f"pytest-warning-no-quality-rule-{uuid4()}")
        db.add(project)
        db.commit()
        project_id = project.id
    finally:
        db.close()

    with TestClient(app) as client:
        response = client.get(f"/api/projects/{project_id}/warning-rules")

    assert response.status_code == 200
    assert all(row["rule_type"] != "low_data_quality" for row in response.json())


def test_empty_warning_records_list_is_stable() -> None:
    db = SessionLocal()
    try:
        project = Project(name="空预警项目")
        db.add(project)
        db.commit()
        project_id = project.id
    finally:
        db.close()

    with TestClient(app) as client:
        response = client.get(f"/api/projects/{project_id}/warnings")

    assert response.status_code == 200
    assert response.json() == []
