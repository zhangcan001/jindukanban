"""导入校验错误清单 Excel 生成。

把 batch.import_validation_issue 行级问题 + raw_import_row 原始数据合并成一份
Excel：一行一条 issue，附带原始整行数据。便于现场人员拿着清单去 Excel 里直接定位
修复后重新上传。

设计要点：
- 不写到 export_dir。错误清单是一次性产物，没必要落盘；直接返回 BytesIO。
- 同一行可能产生多条 issue（比如数值无法解析 + 计划日期倒挂），每条都独立一行。
- raw_data 反序列化失败时不抛错（防御历史脏数据），降级为原始 JSON 串。
"""

from __future__ import annotations

import io
import json
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.models.import_batch import ImportBatch
from app.models.import_validation_issue import ImportValidationIssue
from app.models.raw_import_row import RawImportRow

LEVEL_LABELS = {
    "error": "错误",
    "warning": "警告",
    "info": "提示",
}

HEADER_FILL = PatternFill(start_color="FFD0E4F5", end_color="FFD0E4F5", fill_type="solid")
ERROR_FILL = PatternFill(start_color="FFFADBD8", end_color="FFFADBD8", fill_type="solid")
WARNING_FILL = PatternFill(start_color="FFFCF3CF", end_color="FFFCF3CF", fill_type="solid")


def build_error_report_workbook(db: Session, batch: ImportBatch) -> bytes:
    issues = list(
        db.execute(
            select(ImportValidationIssue)
            .where(ImportValidationIssue.batch_id == batch.id)
            .order_by(
                ImportValidationIssue.row_index.asc().nullsfirst(),
                ImportValidationIssue.id.asc(),
            )
        ).scalars()
    )
    raw_rows_by_index = _load_raw_rows(db, batch.id)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "导入错误清单"

    headers = ["行号", "级别", "错误码", "错误说明", "列名", "原始值", "整行原始数据"]
    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")

    if not issues:
        sheet.append(["", "", "", "无校验问题。", "", "", ""])
    else:
        for issue in issues:
            row_index = issue.row_index
            raw_row = raw_rows_by_index.get(row_index) if row_index is not None else None
            raw_value = ""
            if raw_row is not None and issue.column_name:
                raw_value = _stringify(raw_row.get(issue.column_name, ""))
            row_dump = json.dumps(raw_row, ensure_ascii=False, default=str) if raw_row else ""
            sheet.append(
                [
                    row_index if row_index is not None else "",
                    LEVEL_LABELS.get(issue.level, issue.level or ""),
                    issue.code or "",
                    issue.message or "",
                    issue.column_name or "",
                    raw_value,
                    row_dump,
                ]
            )
            row_cells = sheet[sheet.max_row]
            fill = ERROR_FILL if issue.level == "error" else WARNING_FILL if issue.level == "warning" else None
            if fill is not None:
                for cell in row_cells:
                    cell.fill = fill

    _autosize_columns(sheet, max_width=60)

    summary = workbook.create_sheet("批次信息")
    summary.append(["文件", batch.file_name or ""])
    summary.append(["Sheet", batch.sheet_name or ""])
    summary.append(["数据日期", str(batch.data_date) if batch.data_date else ""])
    summary.append(["导入策略", batch.import_strategy or ""])
    summary.append(["状态", batch.status or ""])
    summary.append(["错误数", batch.error_count or 0])
    summary.append(["警告数", batch.warning_count or 0])
    summary.append(["跳过数", batch.skipped_count or 0])
    summary.append(["导入数", batch.imported_count or 0])
    _autosize_columns(summary, max_width=40)

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _load_raw_rows(db: Session, batch_id: int) -> dict[int, dict[str, Any]]:
    rows: dict[int, dict[str, Any]] = {}
    for raw in db.execute(
        select(RawImportRow).where(RawImportRow.batch_id == batch_id)
    ).scalars():
        try:
            parsed = json.loads(raw.raw_data) if raw.raw_data else {}
        except (TypeError, ValueError):
            parsed = {"_raw": raw.raw_data}
        if isinstance(parsed, dict):
            rows[raw.row_index] = parsed
    return rows


def _stringify(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return value
    return str(value)


def _autosize_columns(sheet, *, max_width: int = 60) -> None:
    for column_cells in sheet.columns:
        length = 8
        for cell in column_cells:
            if cell.value is None:
                continue
            text = str(cell.value)
            length = max(length, min(max_width, _display_width(text)))
        sheet.column_dimensions[column_cells[0].column_letter].width = length + 2


def _display_width(text: str) -> int:
    """中文字符算 2，其它算 1，避免列宽被中文挤窄。"""
    width = 0
    for char in text:
        width += 2 if ord(char) > 0x2E80 else 1
    return width
