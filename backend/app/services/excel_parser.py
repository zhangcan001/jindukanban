from __future__ import annotations

from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook

from app.services.field_detector import detect_column
from app.services.field_diagnostics_service import explain_mapping


class ExcelParseError(ValueError):
    def __init__(self, message: str, code: str = "EXCEL_PARSE_ERROR") -> None:
        super().__init__(message)
        self.code = code


def get_sheet_names(file_path: str) -> list[str]:
    suffix = Path(file_path).suffix.lower()
    if suffix == ".csv":
        return ["CSV"]
    if suffix == ".xlsx":
        workbook = load_workbook(file_path, read_only=True, data_only=True)
        try:
            return list(workbook.sheetnames)
        finally:
            workbook.close()
    raise ExcelParseError("Only .xlsx and .csv files are supported")


HEADER_KEYWORDS = {
    "区域",
    "系统",
    "子项",
    "清单量",
    "累计完成",
    "本期完成",
    "应完成率",
    "进度百分比",
    "楼栋",
    "楼层",
    "专业",
    "工作内容",
    "总工程量",
    "责任人",
    "班组",
    "材料状态",
    "备注",
}


def resolve_header_rows(
    file_path: str,
    sheet_name: str,
    header_row_index: int | None,
    data_start_row_index: int | None,
) -> tuple[int, int]:
    raw = _read_raw(file_path, sheet_name)
    if raw.empty:
        raise ExcelParseError("The selected sheet has no data")

    recommendation = recommend_header_rows(raw)
    resolved_header = header_row_index or recommendation["header_row_index"] or detect_header_row(raw)
    resolved_data_start = data_start_row_index or recommendation["data_start_row_index"] or resolved_header + 1
    return resolved_header, resolved_data_start


def detect_header_row(raw: pd.DataFrame, max_scan_rows: int = 20) -> int:
    return int(recommend_header_rows(raw, max_scan_rows)["header_row_index"] or 1)


def recommend_header_rows(raw: pd.DataFrame, max_scan_rows: int = 20) -> dict[str, int | str | None]:
    best_row = 0
    best_score = float("-inf")
    scan_count = min(max_scan_rows, len(raw.index))
    scored_rows = 0

    for row_index in range(scan_count):
        values = [_clean_cell(value) for value in raw.iloc[row_index].tolist()]
        non_empty = [value for value in values if value]
        if not non_empty:
            continue

        scored_rows += 1
        non_empty_count = len(non_empty)
        keyword_hits = sum(1 for value in non_empty for keyword in HEADER_KEYWORDS if keyword in value)
        detected_hits = sum(1 for value in non_empty if detect_column(value).get("recommended_field"))
        text_count = sum(1 for value in non_empty if not _looks_numeric(value))
        empty_count = len(values) - non_empty_count
        next_non_empty = _next_data_density(raw, row_index)

        score = float(non_empty_count)
        score += keyword_hits * 3
        score += detected_hits * 4
        score += min(text_count, 6) * 0.5
        score += min(next_non_empty, 6) * 0.8
        score -= empty_count * 0.35

        if any("未命名字段" in value for value in non_empty):
            score -= 5
        if non_empty_count == 1 and len(non_empty[0]) >= 18:
            score -= 12
        if non_empty_count <= 2 and empty_count >= max(3, len(values) // 2):
            score -= 4

        if score > best_score:
            best_score = score
            best_row = row_index

    if scored_rows == 0 or best_score < 1:
        return {"header_row_index": None, "data_start_row_index": None, "confidence": "低"}

    header_start, header_end = _adjust_for_parent_header_row(raw, best_row)
    if header_start != best_row:
        confidence = "高" if best_score >= 18 else ("中" if best_score >= 9 else "低")
        return {
            "header_row_index": header_start + 1,
            "data_start_row_index": min(header_end + 2, len(raw.index) + 1),
            "confidence": confidence,
        }

    confidence = "高" if best_score >= 18 else ("中" if best_score >= 9 else "低")
    return {
        "header_row_index": best_row + 1,
        "data_start_row_index": min(best_row + 2, len(raw.index) + 1),
        "confidence": confidence,
    }


def parse_preview(
    file_path: str,
    sheet_name: str,
    header_row_index: int | None,
    data_start_row_index: int | None,
    multi_header: bool,
    header_end_row_index: int | None,
) -> tuple[list[dict[str, str | None]], list[dict[str, Any]], int]:
    raw = _read_raw(file_path, sheet_name)

    if raw.empty:
        raise ExcelParseError("The selected sheet has no data")

    recommendation = recommend_header_rows(raw)
    header_row_index = header_row_index or int(recommendation["header_row_index"] or detect_header_row(raw))
    data_start_row_index = data_start_row_index or int(recommendation["data_start_row_index"] or header_row_index + 1)
    if not multi_header and header_end_row_index is None:
        inferred_header_end = infer_multi_header_end(raw, header_row_index)
        if inferred_header_end is not None:
            multi_header = True
            header_end_row_index = inferred_header_end
            if data_start_row_index <= inferred_header_end:
                data_start_row_index = inferred_header_end + 1
    header_start = header_row_index - 1
    data_start = data_start_row_index - 1
    if header_start >= len(raw.index):
        raise ExcelParseError("header_row_index is outside the file range")
    if data_start > len(raw.index):
        raise ExcelParseError("data_start_row_index is outside the file range")

    if multi_header:
        header_end = (header_end_row_index or header_row_index) - 1
        if header_end < header_start:
            raise ExcelParseError("header_end_row_index must be greater than or equal to header_row_index")
        header_rows = raw.iloc[header_start : header_end + 1]
        normalized_header_rows = _normalize_header_rows(header_rows)
        column_names = [_join_header_values(normalized_header_rows[column_index].tolist()) for column_index in raw.columns]
    else:
        column_names = [_clean_cell(value) or f"未命名字段{index + 1}" for index, value in enumerate(raw.iloc[header_start].tolist())]

    data = raw.iloc[data_start:].copy()
    data.columns = _dedupe_columns(column_names)
    data = data.dropna(how="all")
    preview = data.head(20).where(pd.notna(data), None).to_dict(orient="records")
    columns = []
    for name in data.columns:
        sample_values = _sample_values(data[name].tolist())
        detected = detect_column(str(name), sample_values)
        field_type = detected["field_type"] or "unknown"
        explanation = explain_mapping(str(name), detected["recommended_field"], multi_header=multi_header)
        if detected.get("sample_reason"):
            explanation["reason"] = f"{explanation['reason']} {detected['sample_reason']}"
        if detected.get("needs_review"):
            explanation["confidence"] = "中"
        columns.append(
            {
                "name": str(name),
                "field_type": field_type,
                "recommended_field": detected["recommended_field"],
                "is_dimension": field_type == "text",
                "is_metric": field_type in {"number", "percent", "currency"},
                "save_to_extra": detected["recommended_field"] is None,
                "sample_values": sample_values,
                "needs_review": bool(detected.get("needs_review")),
                **explanation,
            }
        )
    return columns, preview, int(len(data.index))


def parse_rows(
    file_path: str,
    sheet_name: str,
    header_row_index: int,
    data_start_row_index: int,
    multi_header: bool,
    header_end_row_index: int | None,
) -> list[dict[str, Any]]:
    _, preview_rows, _ = parse_preview(
        file_path,
        sheet_name,
        header_row_index,
        data_start_row_index,
        multi_header,
        header_end_row_index,
    )
    suffix = Path(file_path).suffix.lower()
    if suffix == ".csv":
        raw = pd.read_csv(file_path, header=None, dtype=object, keep_default_na=False, encoding="utf-8-sig")
    else:
        raw = _read_raw(file_path, sheet_name)

    header_start = header_row_index - 1
    data_start = data_start_row_index - 1
    if multi_header:
        header_end = (header_end_row_index or header_row_index) - 1
        header_rows = raw.iloc[header_start : header_end + 1]
        normalized_header_rows = _normalize_header_rows(header_rows)
        column_names = [_join_header_values(normalized_header_rows[column_index].tolist()) for column_index in raw.columns]
    else:
        column_names = [_clean_cell(value) or f"未命名字段{index + 1}" for index, value in enumerate(raw.iloc[header_start].tolist())]

    data = raw.iloc[data_start:].copy()
    data.columns = _dedupe_columns(column_names)
    data = data.dropna(how="all")
    return data.where(pd.notna(data), None).to_dict(orient="records")


def infer_multi_header_end(raw: pd.DataFrame, header_row_index: int, max_extra_rows: int = 2) -> int | None:
    header_start = header_row_index - 1
    if header_start < 0 or header_start >= len(raw.index) - 1:
        return None

    single_names = [_clean_cell(value) or f"未命名字段{index + 1}" for index, value in enumerate(raw.iloc[header_start].tolist())]
    single_score = _header_detect_score(single_names)
    best_end: int | None = None
    best_score = single_score
    max_header_end = min(len(raw.index) - 1, header_start + max_extra_rows)

    for header_end in range(header_start + 1, max_header_end + 1):
        candidate_rows = raw.iloc[header_start : header_end + 1]
        normalized_candidate_rows = _normalize_header_rows(candidate_rows)
        candidate_names = [_join_header_values(normalized_candidate_rows[column_index].tolist()) for column_index in raw.columns]
        candidate_score = _header_detect_score(candidate_names)
        next_density = _next_data_density(raw, header_end)
        candidate_label_density = _header_label_density(candidate_rows)
        ending_row_label_density = _header_row_label_density(raw.iloc[header_end].tolist())
        single_duplicates = len(single_names) - len(set(single_names))
        candidate_duplicates = len(candidate_names) - len(set(candidate_names))
        duplicate_improved = candidate_duplicates < single_duplicates
        if (
            candidate_score >= single_score
            and (candidate_score > best_score or duplicate_improved)
            and next_density >= 2
            and candidate_label_density >= 0.35
            and ending_row_label_density >= 0.4
        ):
            best_score = candidate_score
            best_end = header_end + 1

    return best_end


def _join_header_values(values: list[Any]) -> str:
    cleaned = [_clean_cell(value) for value in values]
    parts: list[str] = []
    for value in cleaned:
        if value and value not in parts:
            parts.append(value)
    if len(parts) >= 2 and parts[-1] and parts[-2] in {"工程量", "完成率", "数量"}:
        return parts[-1]
    if len(parts) == 2 and parts[0] in {"计划", "实际"} and (parts[1].endswith("率") or parts[1].endswith("量") or parts[1].endswith("日期")):
        joined = "".join(parts)
    elif len(parts) >= 2 and parts[-2] in {"计划", "实际"} and (parts[-1].endswith("率") or parts[-1].endswith("量") or parts[-1].endswith("日期")):
        joined = "_".join(parts[:-2] + ["".join(parts[-2:])])
    else:
        joined = "_".join(parts)
    return joined or "未命名字段"


def _normalize_header_rows(header_rows: pd.DataFrame) -> pd.DataFrame:
    normalized = header_rows.copy()
    if len(normalized.index) < 2:
        return normalized

    parent_row_count = len(normalized.index) - 1
    for row_position in range(parent_row_count):
        row_label = normalized.index[row_position]
        values = [_clean_cell(value) for value in normalized.iloc[row_position].tolist()]
        non_empty = [value for value in values if value]
        if len(non_empty) < 2:
            continue

        carried = ""
        for column_label, value in zip(normalized.columns, values):
            if value:
                carried = value
            elif carried:
                normalized.at[row_label, column_label] = carried
    return normalized


def _header_detect_score(column_names: list[str]) -> int:
    return sum(1 for name in column_names if detect_column(name).get("recommended_field"))


def _header_label_density(header_rows: pd.DataFrame) -> float:
    values = [_clean_cell(value) for row in header_rows.values.tolist() for value in row]
    non_empty = [value for value in values if value]
    if not values:
        return 0
    label_like = [value for value in non_empty if detect_column(value).get("recommended_field") or any(keyword in value for keyword in HEADER_KEYWORDS)]
    return len(label_like) / max(len(values), 1)


def _header_row_label_density(values: list[Any]) -> float:
    cleaned = [_clean_cell(value) for value in values]
    non_empty = [value for value in cleaned if value]
    if not cleaned:
        return 0
    label_like = [value for value in non_empty if detect_column(value).get("recommended_field") or any(keyword in value for keyword in HEADER_KEYWORDS)]
    numeric_like = [value for value in non_empty if _looks_numeric(value)]
    if numeric_like and len(numeric_like) > max(2, len(non_empty) // 3):
        return 0
    return len(label_like) / max(len(cleaned), 1)


def _adjust_for_parent_header_row(raw: pd.DataFrame, best_row: int) -> tuple[int, int]:
    if best_row <= 0:
        return best_row, best_row

    current_values = [_clean_cell(value) for value in raw.iloc[best_row].tolist()]
    previous_values = [_clean_cell(value) for value in raw.iloc[best_row - 1].tolist()]
    current_detected = sum(1 for value in current_values if detect_column(value).get("recommended_field"))
    previous_detected = sum(1 for value in previous_values if detect_column(value).get("recommended_field"))
    parent_labels = {"计划", "实际", "工程量", "完成率", "数量", "时间", "工程信息", "任务信息"}
    parent_hits = sum(1 for value in previous_values if any(label in value for label in parent_labels))
    repeated_parent_hits = sum(
        1
        for index in range(1, len(previous_values))
        if previous_values[index] and previous_values[index] == previous_values[index - 1]
    )
    combined_names = [_join_header_values(raw.iloc[best_row - 1 : best_row + 1][column_index].tolist()) for column_index in raw.columns]
    combined_detected = sum(1 for name in combined_names if detect_column(name).get("recommended_field"))
    non_empty_current = [value for value in current_values if value]
    duplicate_child_labels = len(non_empty_current) - len(set(non_empty_current))

    if (
        current_detected >= 4
        and (parent_hits >= 2 or repeated_parent_hits >= 2)
        and (combined_detected >= current_detected or duplicate_child_labels > 0)
        and _next_data_density(raw, best_row) >= 2
    ):
        return best_row - 1, best_row
    return best_row, best_row


def _read_raw(file_path: str, sheet_name: str) -> pd.DataFrame:
    suffix = Path(file_path).suffix.lower()
    if suffix == ".csv":
        return pd.read_csv(file_path, header=None, dtype=object, keep_default_na=False, encoding="utf-8-sig")
    if suffix == ".xlsx":
        sheet_names = get_sheet_names(file_path)
        if sheet_name not in sheet_names:
            raise ExcelParseError(f"Sheet not found: {sheet_name}", "SHEET_NOT_FOUND")
        workbook = load_workbook(file_path, read_only=False, data_only=True)
        try:
            sheet = workbook[sheet_name]
            values = [[cell for cell in row] for row in sheet.iter_rows(values_only=True)]
            if not values:
                return pd.DataFrame()
            max_columns = max((len(row) for row in values), default=0)
            for row in values:
                row.extend([None] * (max_columns - len(row)))
            for merged_range in sheet.merged_cells.ranges:
                min_col, min_row, max_col, max_row = merged_range.bounds
                value = sheet.cell(min_row, min_col).value
                for row_index in range(min_row - 1, max_row):
                    while row_index >= len(values):
                        values.append([None] * max_columns)
                    if max_col > max_columns:
                        for row in values:
                            row.extend([None] * (max_col - max_columns))
                        max_columns = max_col
                    for column_index in range(min_col - 1, max_col):
                        values[row_index][column_index] = value
            return pd.DataFrame(values, dtype=object).fillna("")
        finally:
            workbook.close()
    raise ExcelParseError("Only .xlsx and .csv files are supported")


def _clean_cell(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if text.lower() == "nan":
        return ""
    return text


def _looks_numeric(value: str) -> bool:
    try:
        float(value.strip().rstrip("%"))
    except ValueError:
        return False
    return True


def _next_data_density(raw: pd.DataFrame, row_index: int) -> int:
    if row_index + 1 >= len(raw.index):
        return 0
    next_values = [_clean_cell(value) for value in raw.iloc[row_index + 1].tolist()]
    return sum(1 for value in next_values if value)


def _dedupe_columns(columns: list[str]) -> list[str]:
    counts: dict[str, int] = {}
    result: list[str] = []
    for column in columns:
        key = column or "未命名字段"
        counts[key] = counts.get(key, 0) + 1
        result.append(key if counts[key] == 1 else f"{key}_{counts[key]}")
    return result


def _sample_values(values: list[Any], limit: int = 5) -> list[str]:
    samples: list[str] = []
    seen: set[str] = set()
    for value in values:
        if value is None:
            continue
        text = str(value).strip()
        if not text or text.lower() == "nan":
            continue
        if text in seen:
            continue
        samples.append(text)
        seen.add(text)
        if len(samples) >= limit:
            break
    return samples
